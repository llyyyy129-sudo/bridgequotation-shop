from fastapi import FastAPI, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware

from database import engine, SessionLocal
from models import Product, User, Order, PricingSetting, ProformaInvoice, PIHistory, SharedSalesFile, SharedSalesFolder, Base

from reportlab.lib.pagesizes import A4

from io import BytesIO
from datetime import datetime
from decimal import Decimal

from sqlalchemy import text, inspect, func
import json
import base64
import re
import uuid
import shutil
import tempfile
from pathlib import Path


app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")

Base.metadata.create_all(bind=engine)


# =========================
# AUTO DATABASE MIGRATION
# =========================

def column_exists(table_name, column_name):
    inspector = inspect(engine)
    columns = inspector.get_columns(table_name)
    return any(column["name"] == column_name for column in columns)


def add_column_if_missing(table_name, column_name, column_sql):
    try:
        if not column_exists(table_name, column_name):
            with engine.connect() as conn:
                conn.execute(text(f"""
                    ALTER TABLE {table_name}
                    ADD COLUMN {column_name} {column_sql};
                """))
                conn.commit()
                print(f"Added column: {table_name}.{column_name}")
    except Exception as e:
        print(f"Migration skipped for {table_name}.{column_name}:", e)


def run_migrations():
    add_column_if_missing(
        "users",
        "role",
        "VARCHAR DEFAULT 'customer'"
    )

    add_column_if_missing(
        "users",
        "assigned_sales",
        "VARCHAR DEFAULT 'BILL'"
    )

    add_column_if_missing(
        "users",
        "company_name",
        "VARCHAR DEFAULT ''"
    )

    add_column_if_missing(
        "users",
        "email",
        "VARCHAR DEFAULT ''"
    )

    add_column_if_missing(
        "users",
        "account_type",
        "VARCHAR DEFAULT 'customer'"
    )

    add_column_if_missing(
        "users",
        "approval_status",
        "VARCHAR DEFAULT 'Approved'"
    )

    add_column_if_missing(
        "users",
        "customer_level",
        "VARCHAR DEFAULT 'A'"
    )

    add_column_if_missing(
        "products",
        "image_2",
        "VARCHAR DEFAULT ''"
    )

    add_column_if_missing(
        "products",
        "image_3",
        "VARCHAR DEFAULT ''"
    )

    add_column_if_missing(
        "products",
        "image_4",
        "VARCHAR DEFAULT ''"
    )

    add_column_if_missing(
        "products",
        "image_5",
        "VARCHAR DEFAULT ''"
    )

    add_column_if_missing(
        "products",
        "image_6",
        "VARCHAR DEFAULT ''"
    )

    add_column_if_missing(
        "products",
        "video",
        "VARCHAR DEFAULT ''"
    )

    add_column_if_missing(
        "products",
        "packing",
        "VARCHAR DEFAULT ''"
    )

    add_column_if_missing(
        "products",
        "slogan",
        "VARCHAR DEFAULT ''"
    )

    add_column_if_missing(
        "products",
        "is_active",
        "INTEGER DEFAULT 1"
    )

    # Detailed carton / shipping information. These are additive migrations so
    # existing products and the legacy `volume` field remain untouched.
    add_column_if_missing("products", "carton_length", "DOUBLE PRECISION DEFAULT 0")
    add_column_if_missing("products", "carton_width", "DOUBLE PRECISION DEFAULT 0")
    add_column_if_missing("products", "carton_height", "DOUBLE PRECISION DEFAULT 0")
    add_column_if_missing("products", "carton_cbm", "DOUBLE PRECISION DEFAULT 0")
    add_column_if_missing("products", "pcs_per_carton", "INTEGER DEFAULT 0")
    add_column_if_missing("products", "loading_20gp", "INTEGER DEFAULT 0")
    add_column_if_missing("products", "loading_40gp", "INTEGER DEFAULT 0")
    add_column_if_missing("products", "loading_40hq", "INTEGER DEFAULT 0")
    add_column_if_missing("products", "nearest_port", "VARCHAR DEFAULT ''")
    add_column_if_missing("products", "lead_time", "TEXT DEFAULT ''")

    add_column_if_missing(
        "orders",
        "sales_username",
        "VARCHAR DEFAULT 'BILL'"
    )

    add_column_if_missing(
        "orders",
        "status",
        "VARCHAR DEFAULT 'Pending'"
    )

    add_column_if_missing(
        "orders",
        "created_at",
        "VARCHAR DEFAULT ''"
    )

    add_column_if_missing(
        "orders",
        "return_comment",
        "VARCHAR DEFAULT ''"
    )

    add_column_if_missing(
        "shared_sales_files",
        "folder_id",
        "INTEGER DEFAULT 0"
    )


run_migrations()


def migrate_product_price_to_float():
    try:
        if "postgresql" not in str(engine.url):
            return

        inspector = inspect(engine)
        columns = {
            column["name"]: str(column["type"]).lower()
            for column in inspector.get_columns("products")
        }

        price_type = columns.get("price", "")

        if "int" not in price_type:
            return

        with engine.begin() as conn:
            conn.execute(text("""
                ALTER TABLE products
                ALTER COLUMN price TYPE DOUBLE PRECISION
                USING price::double precision
            """))

        print("Migrated products.price to DOUBLE PRECISION.")
    except Exception as e:
        print("Product price type migration skipped:", e)


migrate_product_price_to_float()


# Bootstrap administrators guarantee that the original owner accounts remain
# available after a new deployment. New administrators should be created from
# the Admin Dashboard and are stored in PostgreSQL, so adding them does not
# require another code change or Railway deployment.
BOOTSTRAP_ADMIN_ACCOUNTS = {
    "orange": {
        "password": "Orange123456",
        "company_name": "COPEC"
    },
    "BILLADMIN": {
        "password": "NBFuture2023!",
        "company_name": "COPEC"
    },
    "clover": {
        "password": "12345678",
        "company_name": "COPEC"
    }
}

PROTECTED_BOOTSTRAP_ADMIN_USERNAMES = {
    username.lower()
    for username in BOOTSTRAP_ADMIN_ACCOUNTS
}


def ensure_bootstrap_admin_accounts():
    db = SessionLocal()

    try:
        # Bootstrap accounts are only a recovery mechanism for a brand-new or
        # empty database. Once at least one approved admin exists, usernames are
        # fully database-managed and may be renamed from the Admin Dashboard.
        existing_admin = db.query(User).filter(
            func.lower(User.role) == "admin",
            User.approval_status == "Approved"
        ).first()

        if existing_admin:
            return

        for username, settings in BOOTSTRAP_ADMIN_ACCOUNTS.items():
            admin_user = db.query(User).filter(
                func.lower(User.username) == username.lower()
            ).first()

            if admin_user:
                admin_user.role = "admin"
                admin_user.account_type = "employee"
                admin_user.approval_status = "Approved"
                admin_user.assigned_sales = ""
                admin_user.company_name = (
                    admin_user.company_name or settings["company_name"]
                )
                admin_user.password = admin_user.password or settings["password"]
                continue

            admin_user = User(
                username=username,
                password=settings["password"],
                role="admin",
                assigned_sales="",
                company_name=settings["company_name"],
                email="",
                account_type="employee",
                approval_status="Approved",
                customer_level="A"
            )
            db.add(admin_user)

        db.commit()
    finally:
        db.close()


ensure_bootstrap_admin_accounts()


def ensure_pricing_setting():
    db = SessionLocal()

    setting = db.query(PricingSetting).filter(
        PricingSetting.id == 1
    ).first()

    if not setting:
        setting = PricingSetting(
            id=1,
            b_multiplier=1.2
        )
        db.add(setting)
        db.commit()

    db.close()


ensure_pricing_setting()


def get_b_multiplier(db):
    setting = db.query(PricingSetting).filter(
        PricingSetting.id == 1
    ).first()

    if not setting:
        setting = PricingSetting(
            id=1,
            b_multiplier=1.2
        )
        db.add(setting)
        db.commit()

    return float(setting.b_multiplier or 1.2)


def get_price_multiplier(db, username=None):
    if not username:
        return 1.0

    user = db.query(User).filter(
        User.username == username
    ).first()

    if not user:
        return 1.0

    level = (user.customer_level or "A").upper()

    if level == "B":
        return get_b_multiplier(db)

    return 1.0


def apply_price_multiplier(value, multiplier):
    if value is None:
        return value

    try:
        result = Decimal(str(value)) * Decimal(str(multiplier))
        return float(result)
    except Exception:
        return value




def safe_load_order_items(raw_items):
    if not raw_items:
        return []

    if isinstance(raw_items, list):
        return raw_items

    try:
        loaded = json.loads(raw_items)
        if isinstance(loaded, list):
            return loaded
        return []
    except Exception:
        return []


def normalize_order_items(order):
    items = safe_load_order_items(order.items)

    for item in items:
        if not isinstance(item, dict):
            continue

        current_status = item.get("item_status")

        if not current_status:
            if order.status == "Confirmed":
                current_status = "Confirmed"
            elif order.status in ["Returned", "Rejected"]:
                current_status = "Returned"
            else:
                current_status = "Pending"

            item["item_status"] = current_status

        if "item_return_comment" not in item:
            if current_status == "Returned":
                item["item_return_comment"] = order.return_comment or ""
            else:
                item["item_return_comment"] = ""

    return items


def recalculate_order_status(order, items):
    statuses = []

    for item in items:
        if not isinstance(item, dict):
            continue

        statuses.append(item.get("item_status", "Pending"))

    if not statuses:
        order.status = "Pending"
        order.return_comment = ""
        order.items = json.dumps(items)
        return

    if all(status == "Confirmed" for status in statuses):
        order.status = "Confirmed"
    elif all(status == "Returned" for status in statuses):
        order.status = "Returned"
    elif any(status == "Pending" for status in statuses):
        order.status = "Pending"
    else:
        # Mixed confirmed + returned, no pending products left.
        # The order is finished, but item details still show which products were returned.
        order.status = "Confirmed"

    return_comments = []

    for item in items:
        if not isinstance(item, dict):
            continue

        if item.get("item_status") == "Returned":
            comment = item.get("item_return_comment", "").strip()
            if comment:
                product_name = item.get("name", "Product")
                return_comments.append(f"{product_name}: {comment}")

    order.return_comment = "; ".join(return_comments)
    order.items = json.dumps(items)


def serialize_product(product, multiplier=1.0):
    return {
        "id": product.id,
        "name": product.name,
        "price": apply_price_multiplier(product.price, multiplier),
        "description": product.description,
        "slogan": getattr(product, "slogan", "") or "",
        "image": product.image,
        "image_2": product.image_2,
        "image_3": product.image_3,
        "image_4": product.image_4,
        "image_5": product.image_5,
        "image_6": product.image_6,
        "gallery_images": [
            image for image in [
                product.image,
                product.image_2,
                product.image_3,
                product.image_4,
                product.image_5,
                product.image_6
            ]
            if image and str(image).strip()
        ],
        "video": product.video,
        "packing": product.packing,
        "moq": product.moq,
        "material": product.material,
        "volume": product.volume,
        "carton_length": getattr(product, "carton_length", 0.0) or 0.0,
        "carton_width": getattr(product, "carton_width", 0.0) or 0.0,
        "carton_height": getattr(product, "carton_height", 0.0) or 0.0,
        "carton_cbm": getattr(product, "carton_cbm", 0.0) or 0.0,
        "pcs_per_carton": getattr(product, "pcs_per_carton", 0) or 0,
        "loading_20gp": getattr(product, "loading_20gp", 0) or 0,
        "loading_40gp": getattr(product, "loading_40gp", 0) or 0,
        "loading_40hq": getattr(product, "loading_40hq", 0) or 0,
        "nearest_port": getattr(product, "nearest_port", "") or "",
        "lead_time": getattr(product, "lead_time", "") or "",
        "size": product.size,
        "price_500": apply_price_multiplier(product.price_500, multiplier),
        "price_1000": apply_price_multiplier(product.price_1000, multiplier),
        "price_3000": apply_price_multiplier(product.price_3000, multiplier),
        "price_10000": apply_price_multiplier(product.price_10000, multiplier),
        "price_50000": apply_price_multiplier(product.price_50000, multiplier),
        "category": product.category,
        "is_active": getattr(product, "is_active", 1)
    }



# =========================
# PROFORMA INVOICE HELPERS
# =========================

def order_all_items_confirmed(order):
    items = normalize_order_items(order)

    if not items:
        return False

    for item in items:
        if not isinstance(item, dict):
            return False

        if item.get("item_status") != "Confirmed":
            return False

    return True


def get_latest_pi(db, order_id):
    return db.query(ProformaInvoice).filter(
        ProformaInvoice.order_id == order_id
    ).order_by(ProformaInvoice.id.desc()).first()


def pi_waiting_for_customer(pi):
    if not pi:
        return False

    return str(pi.status or "").strip().lower() in {
        "sent",
        "revised sent"
    }


def serialize_pi(pi):
    if not pi:
        return None

    return {
        "id": pi.id,
        "order_id": pi.order_id,
        "pi_no": pi.pi_no,
        "pdf_path": pi.pdf_path,
        "pdf_url": pi.pdf_path,
        "status": pi.status,
        "version": pi.version,
        "sent_by": pi.sent_by,
        "sent_at": pi.sent_at,
        "received_at": pi.received_at,
        "customer_message": pi.customer_message or ""
    }


def serialize_pi_history_item(history):
    return {
        "id": history.id,
        "order_id": history.order_id,
        "pi_id": history.pi_id,
        "action": history.action,
        "message": history.message or "",
        "created_by": history.created_by or "",
        "created_at": history.created_at or "",
        "pdf_path": history.pdf_path or ""
    }


def serialize_pi_state(db, order):
    current_pi = get_latest_pi(db, order.id)

    histories = db.query(PIHistory).filter(
        PIHistory.order_id == order.id
    ).order_by(PIHistory.id.asc()).all()

    return {
        "has_pi": current_pi is not None,
        "current": serialize_pi(current_pi),
        "history": [
            serialize_pi_history_item(history)
            for history in histories
        ]
    }


def add_pi_history(db, order_id, pi_id, action, message, created_by, pdf_path=""):
    history = PIHistory(
        order_id=order_id,
        pi_id=pi_id,
        action=action,
        message=message or "",
        created_by=created_by or "",
        created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        pdf_path=pdf_path or ""
    )

    db.add(history)
    return history


def generate_pi_pdf_file(order, customer, sales_user, pi_no, version):
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        Image
    )
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.utils import ImageReader
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from pathlib import Path as LocalPath

    pi_dir = LocalPath("static/uploads/pi")
    pi_dir.mkdir(parents=True, exist_ok=True)

    safe_pi_no = re.sub(r"[^A-Za-z0-9_-]+", "_", pi_no)
    file_name = f"{safe_pi_no}_v{version}.pdf"
    output_path = pi_dir / file_name

    styles = getSampleStyleSheet()

    brand_blue = colors.HexColor("#1f3c88")
    light_blue = colors.HexColor("#eef4ff")
    soft_gray = colors.HexColor("#f8fafc")
    border_gray = colors.HexColor("#dbe3ef")
    dark_text = colors.HexColor("#111827")
    muted_text = colors.HexColor("#4b5563")

    title_style = ParagraphStyle(
        "PITitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        textColor=brand_blue,
        alignment=TA_CENTER,
        spaceAfter=8
    )

    normal_style = ParagraphStyle(
        "PINormal",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        textColor=dark_text
    )

    small_style = ParagraphStyle(
        "PISmall",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
        textColor=dark_text
    )

    small_center_style = ParagraphStyle(
        "PISmallCenter",
        parent=small_style,
        alignment=TA_CENTER
    )

    small_right_style = ParagraphStyle(
        "PISmallRight",
        parent=small_style,
        alignment=TA_RIGHT
    )

    def escape_pdf_text(value):
        if value is None:
            value = ""

        return (
            str(value)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace("\n", "<br/>")
        )

    def p(value, style=small_style):
        return Paragraph(escape_pdf_text(value), style)

    def money(value):
        try:
            return "${:,.2f}".format(float(value or 0))
        except Exception:
            return "$0.00"

    def safe_number(value, default=0):
        try:
            return float(value or default)
        except Exception:
            return float(default)

    def make_product_image(image_url):
        if not image_url:
            return p("No Image", small_center_style)

        image_path = str(image_url)

        if image_path.startswith("/"):
            image_path = "." + image_path

        local_path = LocalPath(image_path)

        if not local_path.exists():
            return p("No Image", small_center_style)

        try:
            reader = ImageReader(str(local_path))
            img_width, img_height = reader.getSize()

            max_width = 52
            max_height = 48
            scale = min(max_width / img_width, max_height / img_height)

            product_image = Image(
                str(local_path),
                width=img_width * scale,
                height=img_height * scale
            )

            product_image.hAlign = "CENTER"
            return product_image

        except Exception:
            return p("No Image", small_center_style)

    items = normalize_order_items(order)

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=20,
        leftMargin=20,
        topMargin=18,
        bottomMargin=18
    )

    elements = []

    elements.append(Paragraph("PROFORMA INVOICE / SALES CONFIRMATION", title_style))

    header_data = [
        [
            Paragraph(
                f"<b>S/C NO.:</b> {escape_pdf_text(pi_no)}",
                normal_style
            ),
            Paragraph(
                f"<b>DATE:</b> {datetime.now().strftime('%b %d, %Y')}",
                normal_style
            )
        ]
    ]

    header_table = Table(header_data, colWidths=[360, 175])
    header_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.7, border_gray),
        ("BACKGROUND", (0, 0), (-1, -1), light_blue),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 8))

    buyer_company = customer.company_name if customer else order.username
    buyer_email = customer.email if customer else ""
    buyer_contact = customer.username if customer else order.username

    seller_company = "NINGBO FUTURE HOUSEWARE CO.,LTD."
    seller_address = "2560 YONGJIANG AVENUE, BUILDING 8 IN WEST DISTRICT, NINGBO 315048, CHINA"
    seller_contact = sales_user.username if sales_user else (order.sales_username or "")
    seller_email = sales_user.email if sales_user else ""

    buyer_block = (
        "<b>THE BUYER:</b><br/>"
        + escape_pdf_text(buyer_company or "")
        + "<br/>CONTACT: "
        + escape_pdf_text(buyer_contact or "")
        + "<br/>EMAIL: "
        + escape_pdf_text(buyer_email or "")
    )

    seller_block = (
        "<b>THE SELLER:</b><br/>"
        + escape_pdf_text(seller_company)
        + "<br/>"
        + escape_pdf_text(seller_address)
        + "<br/>CONTACT: "
        + escape_pdf_text(seller_contact or "")
        + "<br/>EMAIL: "
        + escape_pdf_text(seller_email or "")
    )

    parties_table = Table(
        [[Paragraph(buyer_block, normal_style), Paragraph(seller_block, normal_style)]],
        colWidths=[267, 267]
    )
    parties_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.7, border_gray),
        ("INNERGRID", (0, 0), (-1, -1), 0.6, border_gray),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(parties_table)
    elements.append(Spacer(1, 9))

    intro = Paragraph(
        "This undersigned seller and buyer have agreed to conclude the following transactions according to the terms and conditions stipulated below:",
        normal_style
    )
    elements.append(intro)
    elements.append(Spacer(1, 8))

    table_data = [
        [
            p("Picture", small_center_style),
            p("Description", small_center_style),
            p("Ctn", small_center_style),
            p("Inner", small_center_style),
            p("Unit", small_center_style),
            p("Qty", small_center_style),
            p("Meas./ctn", small_center_style),
            p("Meas.", small_center_style),
            Paragraph("Price<br/>FOB NINGBO", small_center_style),
            p("Amount", small_center_style)
        ]
    ]

    total_qty = 0
    grand_total = 0

    for item in items:
        qty = int(safe_number(item.get("quantity") or item.get("qty"), 0))
        unit_price = safe_number(item.get("price") or item.get("unit_price"), 0)
        amount = qty * unit_price

        total_qty += qty
        grand_total += amount

        description_parts = [
            item.get("name") or item.get("product_name") or "Product",
            f"Size: {item.get('size', '')}",
            f"Material: {item.get('material', '')}",
            f"Packaging: {item.get('packing', '')}",
        ]

        if item.get("customer_requirement"):
            description_parts.append(f"Requirement: {item.get('customer_requirement', '')}")

        description = "\n".join([part for part in description_parts if part])

        volume = item.get("volume", "")

        table_data.append([
            make_product_image(item.get("image", "")),
            p(description),
            p("", small_center_style),
            p("", small_center_style),
            p("PCS", small_center_style),
            p(f"{qty:,}", small_center_style),
            p(volume, small_center_style),
            p("", small_center_style),
            p(money(unit_price), small_right_style),
            p(money(amount), small_right_style)
        ])

    table_data.append([
        "",
        p("TTL:", small_right_style),
        "",
        "",
        "",
        p(f"{total_qty:,}", small_center_style),
        "",
        "",
        "",
        p(money(grand_total), small_right_style)
    ])

    product_table = Table(
        table_data,
        colWidths=[56, 145, 34, 34, 38, 44, 54, 44, 62, 62],
        repeatRows=1
    )

    product_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand_blue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), light_blue),
        ("BOX", (0, 0), (-1, -1), 0.7, border_gray),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, border_gray),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (2, 1), (7, -1), "CENTER"),
        ("ALIGN", (8, 1), (9, -1), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    elements.append(product_table)
    elements.append(Spacer(1, 10))

    terms = [
        ["PORT OF LOADING:", "NINGBO, CHINA"],
        ["PORT OF DESTINATION:", "TO BE CONFIRMED"],
        ["SHIPMENT TERM:", "BY SEA"],
        ["PAYMENT TERM:", "TO BE CONFIRMED"],
        ["TIME OF SHIPMENT:", "TO BE CONFIRMED"]
    ]

    terms_table = Table(
        [[p(label, normal_style), p(value, normal_style)] for label, value in terms],
        colWidths=[150, 385]
    )
    terms_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.7, border_gray),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, border_gray),
        ("BACKGROUND", (0, 0), (0, -1), soft_gray),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(terms_table)
    elements.append(Spacer(1, 10))

    bank_lines = (
        "<b>OUR BANK INFORMATION:</b><br/>"
        "BENEFICIARY: NINGBO FUTURE HOUSEWARE CO.,LTD.<br/>"
        "ADD: 16/F,95 Business Mansion, No.598 Jiangnan Road, Jiangdong, Ningbo<br/>"
        "BENEFICIARY BANK: SHANGHAI PUDONG DEVELOPMENT BANK NINGBO BRANCH<br/>"
        "ADD: 21 JIANG XIA STREET, NINGBO, P.R.CHINA<br/>"
        "SWIFT CODE: SPDBCNSH342<br/>"
        "ACCOUNT NO.: 94171455350000054<br/>"
        "T/T PAYMENT PLEASE CHOOSE INTERMEDIARY BANK: CITIBANK N.A., NEW YORK<br/>"
        "SWIFT CODE: CITIUS33"
    )

    bank_table = Table([[Paragraph(bank_lines, normal_style)]], colWidths=[535])
    bank_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.7, border_gray),
        ("BACKGROUND", (0, 0), (-1, -1), soft_gray),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    elements.append(bank_table)
    elements.append(Spacer(1, 9))

    note = Paragraph(
        "Our company promises that our exported products are fully in line with the quality testing standards of the host country, and we will take full responsibility for any quality problems.<br/>"
        "Please note that we will not change bank details by email. Any account change will be notified by our company in official documents.",
        small_style
    )
    elements.append(note)

    doc.build(elements)

    return "/static/uploads/pi/" + file_name




# Product Excel imports must be run manually from seed.py.
# Do not seed products during application startup, otherwise products deleted
# by an administrator are recreated whenever Railway redeploys the service.


def sync_product_id_sequence():
    try:
        if "postgresql" not in str(engine.url):
            return

        with engine.connect() as conn:
            conn.execute(text("""
                SELECT setval(
                    pg_get_serial_sequence('products', 'id'),
                    COALESCE((SELECT MAX(id) FROM products), 1),
                    true
                );
            """))
            conn.commit()
            print("Product id sequence synced.")
    except Exception as e:
        print("Product id sequence sync skipped:", e)


sync_product_id_sequence()


app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# =========================
# PAGE ROUTES
# =========================

@app.get("/")
def root():
    return FileResponse("templates/login.html")


@app.get("/login.html")
def login_page():
    return FileResponse("templates/login.html")


@app.get("/register.html")
def register_page():
    return FileResponse("templates/register.html")


@app.get("/change_password.html")
def change_password_page():
    return FileResponse("templates/change_password.html")


@app.get("/account_status.html")
def account_status_page():
    return FileResponse("templates/account_status.html")


@app.get("/products.html")
def products_page():
    return FileResponse("templates/products.html")


@app.get("/product.html")
def product_page():
    return FileResponse("templates/product.html")


@app.get("/cart.html")
def cart_page():
    return FileResponse("templates/cart.html")


@app.get("/my_orders.html")
def my_orders_page():
    return FileResponse("templates/my_orders.html")


@app.get("/share.html")
def share_page():
    return FileResponse("templates/share.html")


@app.get("/sales.html")
def sales_page():
    return FileResponse("templates/sales.html")


@app.get("/admin.html")
def admin_page():
    return FileResponse("templates/admin.html")


# =========================
# PRODUCTS
# =========================

@app.get("/products")
def get_products(username: str = ""):
    db = SessionLocal()

    multiplier = get_price_multiplier(db, username)
    products = db.query(Product).filter(
        Product.is_active != 0
    ).all()

    result = [
        serialize_product(product, multiplier)
        for product in products
    ]

    db.close()
    return result


@app.get("/products/{product_id}")
def get_product(product_id: int, username: str = ""):
    db = SessionLocal()

    multiplier = get_price_multiplier(db, username)

    product = db.query(Product).filter(
        Product.id == product_id
    ).first()

    if not product:
        db.close()
        return {
            "error": "Product not found"
        }

    result = serialize_product(product, multiplier)

    db.close()
    return result




# =========================
# AUTH
# =========================

@app.post("/register")
def register(user: dict):
    db = SessionLocal()

    username = user.get("username", "").strip()
    password = user.get("password", "").strip()
    company_name = user.get("company_name", "").strip()
    email = user.get("email", "").strip()
    account_type = user.get("account_type", "customer").strip()

    if not username or not password or not company_name or not email:
        db.close()
        return {
            "success": False,
            "message": "Please fill in username, password, company name and email."
        }

    if len(password) < 8:
        db.close()
        return {
            "success": False,
            "message": "Password must be at least 8 characters."
        }

    if account_type not in ["customer", "employee"]:
        db.close()
        return {
            "success": False,
            "message": "Invalid account type."
        }

    existing_username = db.query(User).filter(
        User.username == username
    ).first()

    if existing_username:
        db.close()
        return {
            "success": False,
            "message": "Username already exists!"
        }

    existing_email = db.query(User).filter(
        User.email == email
    ).first()

    if existing_email:
        db.close()
        return {
            "success": False,
            "message": "Email already exists!"
        }

    if account_type == "employee":
        role = "sales"
        assigned_sales = ""
    else:
        role = "customer"
        assigned_sales = "BILL"

    new_user = User(
        username=username,
        password=password,
        role=role,
        assigned_sales=assigned_sales,
        company_name=company_name,
        email=email,
        account_type=account_type,
        approval_status="Pending",
        customer_level="A"
    )

    db.add(new_user)
    db.commit()
    db.close()

    return {
        "success": True,
        "message": "Registration submitted. Please wait for admin approval."
    }


@app.post("/login")
def login(user: dict):
    db = SessionLocal()

    username = user.get("username", "").strip()
    password = user.get("password", "").strip()

    existing = db.query(User).filter(
        User.username == username,
        User.password == password
    ).first()

    if not existing:
        db.close()
        return {
            "success": False,
            "message": "Invalid username or password"
        }

    approval_status = existing.approval_status or "Approved"

    if approval_status == "Pending":
        result = {
            "success": False,
            "message": "Your account is waiting for admin approval.",
            "username": existing.username,
            "role": existing.role,
            "account_type": existing.account_type,
            "approval_status": "Pending",
            "company_name": existing.company_name,
            "email": existing.email
        }
        db.close()
        return result

    if approval_status == "Rejected":
        result = {
            "success": False,
            "message": "Your registration request has been rejected.",
            "username": existing.username,
            "role": existing.role,
            "account_type": existing.account_type,
            "approval_status": "Rejected",
            "company_name": existing.company_name,
            "email": existing.email
        }
        db.close()
        return result

    result = {
        "success": True,
        "message": "Login successful!",
        "username": existing.username,
        "role": existing.role,
        "assigned_sales": existing.assigned_sales,
        "account_type": existing.account_type,
        "approval_status": approval_status,
        "company_name": existing.company_name,
        "email": existing.email,
        "customer_level": existing.customer_level
    }

    db.close()
    return result


@app.post("/change-password")
def change_password(data: dict):
    db = SessionLocal()

    username = data.get("username", "").strip()
    current_password = data.get("current_password", "").strip()
    new_password = data.get("new_password", "").strip()

    if not username or not current_password or not new_password:
        db.close()
        return {
            "success": False,
            "message": "Please fill in username, current password and new password."
        }

    if len(new_password) < 6:
        db.close()
        return {
            "success": False,
            "message": "New password must be at least 6 characters."
        }

    user = db.query(User).filter(
        User.username == username,
        User.password == current_password
    ).first()

    if not user:
        db.close()
        return {
            "success": False,
            "message": "Invalid username or current password."
        }

    user.password = new_password

    db.commit()
    db.close()

    return {
        "success": True,
        "message": "Password updated successfully. Please log in again."
    }


# =========================
# ADMIN USER APPROVAL
# =========================

def serialize_user(u):
    normalized_username = str(u.username or "").strip().lower()

    return {
        "id": u.id,
        "username": u.username,
        "company_name": u.company_name,
        "email": u.email,
        "account_type": u.account_type,
        "role": u.role,
        "assigned_sales": u.assigned_sales,
        "approval_status": u.approval_status,
        "customer_level": u.customer_level,
        "is_protected_admin": False
    }


def get_user_by_username_case_insensitive(db, username):
    normalized_username = str(username or "").strip().lower()

    if not normalized_username:
        return None

    return db.query(User).filter(
        func.lower(User.username) == normalized_username
    ).first()


def can_approved_admin(db, requester_username):
    admin_user = get_user_by_username_case_insensitive(
        db,
        requester_username
    )

    return bool(
        admin_user and
        str(admin_user.role or "").strip().lower() == "admin" and
        str(admin_user.approval_status or "Approved").strip() == "Approved"
    )


def can_admin_manage_user_passwords(db, requester_username):
    return can_approved_admin(db, requester_username)


@app.post("/admin/users/create-admin")
def create_admin_account(data: dict):
    db = SessionLocal()

    try:
        requester_username = str(
            data.get("requester_username", "") or ""
        ).strip()

        if not can_approved_admin(db, requester_username):
            return {
                "success": False,
                "message": "Only an approved admin can create another admin account."
            }

        username = str(data.get("username", "") or "").strip()
        password = str(data.get("password", "") or "").strip()
        company_name = str(
            data.get("company_name", "COPEC") or "COPEC"
        ).strip()
        email = str(data.get("email", "") or "").strip()

        if not username or not password:
            return {
                "success": False,
                "message": "Username and password are required."
            }

        if not re.fullmatch(r"[A-Za-z0-9_.-]{3,50}", username):
            return {
                "success": False,
                "message": "Username must be 3-50 characters and can only contain letters, numbers, dots, underscores or hyphens."
            }

        if len(password) < 8:
            return {
                "success": False,
                "message": "Admin password must be at least 8 characters."
            }

        existing_username = get_user_by_username_case_insensitive(
            db,
            username
        )

        if existing_username:
            return {
                "success": False,
                "message": "Username already exists."
            }

        if email:
            existing_email = db.query(User).filter(
                func.lower(User.email) == email.lower()
            ).first()

            if existing_email:
                return {
                    "success": False,
                    "message": "Email already exists."
                }

        new_admin = User(
            username=username,
            password=password,
            role="admin",
            assigned_sales="",
            company_name=company_name or "COPEC",
            email=email,
            account_type="employee",
            approval_status="Approved",
            customer_level="A"
        )

        db.add(new_admin)
        db.commit()
        db.refresh(new_admin)

        return {
            "success": True,
            "message": f"Admin account {new_admin.username} has been created.",
            "user": serialize_user(new_admin)
        }

    except Exception as e:
        db.rollback()
        return {
            "success": False,
            "message": "Create admin account failed: " + str(e)
        }
    finally:
        db.close()


@app.get("/admin/users")
def get_all_users():
    db = SessionLocal()

    users = db.query(User).order_by(User.id.desc()).all()
    result = [serialize_user(u) for u in users]

    db.close()
    return result


@app.get("/admin/users/pending")
def get_pending_users():
    db = SessionLocal()

    users = db.query(User).filter(
        User.approval_status == "Pending"
    ).order_by(User.id.desc()).all()

    result = [serialize_user(u) for u in users]

    db.close()
    return result


@app.post("/admin/users/{user_id}/approve")
def approve_user(user_id: int):
    db = SessionLocal()

    user = db.query(User).filter(
        User.id == user_id
    ).first()

    if not user:
        db.close()
        return {
            "success": False,
            "message": "User not found."
        }

    username = user.username

    if str(user.role or "").strip().lower() == "admin":
        db.close()
        return {
            "success": False,
            "message": "Full admin accounts are already approved and cannot be changed."
        }

    user.approval_status = "Approved"

    if user.account_type == "employee":
        user.role = "sales"
        user.assigned_sales = ""
    else:
        user.role = "customer"
        user.assigned_sales = "BILL"
        user.customer_level = user.customer_level or "A"

    db.commit()
    db.close()

    return {
        "success": True,
        "message": f"{username} has been approved."
    }


@app.post("/admin/users/{user_id}/reject")
def reject_user(user_id: int):
    db = SessionLocal()

    user = db.query(User).filter(
        User.id == user_id
    ).first()

    if not user:
        db.close()
        return {
            "success": False,
            "message": "User not found."
        }

    username = user.username

    if str(user.role or "").strip().lower() == "admin":
        db.close()
        return {
            "success": False,
            "message": "Admin user cannot be rejected."
        }

    user.approval_status = "Rejected"

    db.commit()
    db.close()

    return {
        "success": True,
        "message": f"{username} has been rejected."
    }




@app.get("/admin/sales-users")
def get_sales_users():
    db = SessionLocal()

    users = db.query(User).filter(
        User.role == "sales",
        User.approval_status == "Approved"
    ).order_by(User.username.asc()).all()

    result = [
        {
            "id": u.id,
            "username": u.username,
            "company_name": u.company_name,
            "email": u.email
        }
        for u in users
    ]

    db.close()
    return result



@app.get("/admin/pricing")
def get_admin_pricing():
    db = SessionLocal()

    setting = db.query(PricingSetting).filter(
        PricingSetting.id == 1
    ).first()

    if not setting:
        setting = PricingSetting(
            id=1,
            b_multiplier=1.2
        )
        db.add(setting)
        db.commit()

    result = {
        "success": True,
        "b_multiplier": float(setting.b_multiplier or 1.2),
        "b_percentage": round(float(setting.b_multiplier or 1.2) * 100, 2)
    }

    db.close()
    return result


@app.post("/admin/pricing")
def update_admin_pricing(data: dict):
    db = SessionLocal()

    try:
        percentage = float(data.get("b_percentage", 120))
    except Exception:
        percentage = 120

    if percentage < 100:
        db.close()
        return {
            "success": False,
            "message": "B Class percentage cannot be lower than 100%."
        }

    if percentage > 300:
        db.close()
        return {
            "success": False,
            "message": "B Class percentage cannot be higher than 300%."
        }

    setting = db.query(PricingSetting).filter(
        PricingSetting.id == 1
    ).first()

    if not setting:
        setting = PricingSetting(
            id=1,
            b_multiplier=percentage / 100
        )
        db.add(setting)
    else:
        setting.b_multiplier = percentage / 100

    db.commit()
    db.close()

    return {
        "success": True,
        "message": f"B Class price percentage updated to {percentage}%."
    }


@app.post("/admin/customers/{customer_id}/level")
def update_customer_level(customer_id: int, data: dict):
    db = SessionLocal()

    customer = db.query(User).filter(
        User.id == customer_id
    ).first()

    if not customer:
        db.close()
        return {
            "success": False,
            "message": "Customer not found."
        }

    if customer.role != "customer":
        db.close()
        return {
            "success": False,
            "message": "Only customer users can have customer level."
        }

    level = data.get("customer_level", "A").strip().upper()

    if level not in ["A", "B"]:
        db.close()
        return {
            "success": False,
            "message": "Customer level must be A or B."
        }

    customer_username = customer.username
    customer.customer_level = level

    db.commit()
    db.close()

    return {
        "success": True,
        "message": f"{customer_username} customer level updated to {level}."
    }

@app.post("/admin/customers/{customer_id}/assign-sales")
def assign_sales_to_customer(customer_id: int, data: dict):
    db = SessionLocal()

    customer = db.query(User).filter(
        User.id == customer_id
    ).first()

    if not customer:
        db.close()
        return {
            "success": False,
            "message": "Customer not found."
        }

    if customer.role != "customer":
        db.close()
        return {
            "success": False,
            "message": "Only customer users can be assigned to sales."
        }

    sales_username = data.get("sales_username", "").strip()

    if not sales_username:
        db.close()
        return {
            "success": False,
            "message": "Please choose a sales user."
        }

    sales_user = db.query(User).filter(
        User.username == sales_username,
        User.role == "sales",
        User.approval_status == "Approved"
    ).first()

    if not sales_user:
        db.close()
        return {
            "success": False,
            "message": "Sales user not found or not approved."
        }

    customer_username = customer.username

    customer.assigned_sales = sales_username

    # Demo-friendly behavior:
    # move this customer's existing orders to the newly assigned sales.
    orders = db.query(Order).filter(
        Order.username == customer_username
    ).all()

    for order in orders:
        order.sales_username = sales_username

    db.commit()
    db.close()

    return {
        "success": True,
        "message": f"{customer_username} has been assigned to {sales_username}."
    }

@app.post("/admin/users/{user_id}/password-info")
def get_user_password_info(user_id: int, data: dict):
    db = SessionLocal()
    requester_username = str(data.get("requester_username", "") or "").strip()

    if not can_admin_manage_user_passwords(db, requester_username):
        db.close()
        return {
            "success": False,
            "message": "Only an approved admin can view account passwords."
        }

    user = db.query(User).filter(
        User.id == user_id
    ).first()

    if not user:
        db.close()
        return {
            "success": False,
            "message": "User not found."
        }

    result = {
        "success": True,
        "username": user.username,
        "password": user.password or ""
    }
    db.close()
    return result


@app.post("/admin/users/{user_id}/change-password")
def change_user_password(user_id: int, data: dict):
    db = SessionLocal()
    requester_username = str(data.get("requester_username", "") or "").strip()

    if not can_admin_manage_user_passwords(db, requester_username):
        db.close()
        return {
            "success": False,
            "message": "Only an approved admin can reset account passwords."
        }

    user = db.query(User).filter(
        User.id == user_id
    ).first()

    if not user:
        db.close()
        return {
            "success": False,
            "message": "User not found."
        }

    username = user.username
    new_password = data.get("password", "").strip()

    if not new_password:
        db.close()
        return {
            "success": False,
            "message": "Password cannot be empty."
        }

    if len(new_password) < 6:
        db.close()
        return {
            "success": False,
            "message": "Password must be at least 6 characters."
        }

    user.password = new_password

    db.commit()
    db.close()

    return {
        "success": True,
        "message": f"Password for {username} has been updated."
    }

@app.post("/admin/users/{user_id}/delete")
def delete_user(user_id: int, data: dict):
    db = SessionLocal()

    try:
        requester_username = str(
            data.get("requester_username", "") or ""
        ).strip()

        if not can_approved_admin(db, requester_username):
            return {
                "success": False,
                "message": "Only an approved admin can delete accounts."
            }

        user = db.query(User).filter(
            User.id == user_id
        ).first()

        if not user:
            return {
                "success": False,
                "message": "User not found."
            }

        username = user.username
        normalized_username = str(username or "").strip().lower()
        normalized_requester = requester_username.lower()

        if normalized_username == normalized_requester:
            return {
                "success": False,
                "message": "You cannot delete the admin account currently in use."
            }

        if str(user.role or "").strip().lower() == "admin":
            approved_admin_count = db.query(User).filter(
                func.lower(User.role) == "admin",
                User.approval_status == "Approved"
            ).count()

            if approved_admin_count <= 1:
                return {
                    "success": False,
                    "message": "The last approved admin account cannot be deleted."
                }

        db.delete(user)
        db.commit()

        return {
            "success": True,
            "message": f"{username} has been deleted."
        }

    except Exception as e:
        db.rollback()
        return {
            "success": False,
            "message": "Delete account failed: " + str(e)
        }
    finally:
        db.close()



@app.post("/admin/users/{user_id}/profile")
def update_user_profile(user_id: int, data: dict):
    db = SessionLocal()

    try:
        requester_username = str(
            data.get("requester_username", "") or ""
        ).strip()

        if not can_approved_admin(db, requester_username):
            return {
                "success": False,
                "message": "Only an approved admin can edit account profiles or usernames."
            }

        user = db.query(User).filter(
            User.id == user_id
        ).first()

        if not user:
            return {
                "success": False,
                "message": "User not found."
            }

        old_username = str(user.username or "").strip()
        new_username = str(
            data.get("username", old_username) or old_username
        ).strip()
        company_name = str(data.get("company_name", "") or "").strip()
        email = str(data.get("email", "") or "").strip()

        if not re.fullmatch(r"[A-Za-z0-9_.-]{3,50}", new_username):
            return {
                "success": False,
                "message": "Username must be 3-50 characters and can only contain letters, numbers, dots, underscores or hyphens."
            }

        existing_username = db.query(User).filter(
            func.lower(User.username) == new_username.lower(),
            User.id != user_id
        ).first()

        if existing_username:
            return {
                "success": False,
                "message": "Username already exists."
            }

        renamed = new_username != old_username

        if renamed:
            old_username_lower = old_username.lower()

            # Keep all username-based relationships valid after a rename.
            db.query(User).filter(
                func.lower(User.assigned_sales) == old_username_lower
            ).update(
                {User.assigned_sales: new_username},
                synchronize_session=False
            )

            db.query(Order).filter(
                func.lower(Order.username) == old_username_lower
            ).update(
                {Order.username: new_username},
                synchronize_session=False
            )

            db.query(Order).filter(
                func.lower(Order.sales_username) == old_username_lower
            ).update(
                {Order.sales_username: new_username},
                synchronize_session=False
            )

            db.query(ProformaInvoice).filter(
                func.lower(ProformaInvoice.sent_by) == old_username_lower
            ).update(
                {ProformaInvoice.sent_by: new_username},
                synchronize_session=False
            )

            db.query(PIHistory).filter(
                func.lower(PIHistory.created_by) == old_username_lower
            ).update(
                {PIHistory.created_by: new_username},
                synchronize_session=False
            )

            db.query(SharedSalesFile).filter(
                func.lower(SharedSalesFile.uploaded_by) == old_username_lower
            ).update(
                {SharedSalesFile.uploaded_by: new_username},
                synchronize_session=False
            )

            db.query(SharedSalesFolder).filter(
                func.lower(SharedSalesFolder.created_by) == old_username_lower
            ).update(
                {SharedSalesFolder.created_by: new_username},
                synchronize_session=False
            )

            user.username = new_username

        user.company_name = company_name
        user.email = email

        db.commit()
        db.refresh(user)

        return {
            "success": True,
            "message": (
                f"Username changed from {old_username} to {user.username}."
                if renamed
                else f"Profile for {user.username} has been updated."
            ),
            "old_username": old_username,
            "username": user.username,
            "renamed": renamed,
            "user": serialize_user(user)
        }

    except Exception as e:
        db.rollback()
        return {
            "success": False,
            "message": "Profile update failed: " + str(e)
        }
    finally:
        db.close()


# =========================
# ADMIN PRODUCT MANAGEMENT
# =========================

def to_int(value, default=0):
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except Exception:
        return default


def to_float(value, default=0.0):
    try:
        if value is None or value == "":
            return default
        return float(value)
    except Exception:
        return default


# COPEC quotation sheets use these conservative usable-volume assumptions.
# Container loading quantities are stored/displayed as PCS, not cartons.
CONTAINER_USABLE_CBM = {
    "20gp": 26.0,
    "40gp": 56.0,
    "40hq": 66.0,
}


def calculate_carton_cbm(length_cm, width_cm, height_cm):
    if length_cm <= 0 or width_cm <= 0 or height_cm <= 0:
        return 0.0

    return round((length_cm * width_cm * height_cm) / 1_000_000, 6)


def calculate_container_loading_pcs(carton_cbm, usable_cbm, pcs_per_carton):
    if carton_cbm <= 0 or pcs_per_carton <= 0:
        return 0

    carton_count = max(0, int(float(usable_cbm) / float(carton_cbm)))
    return carton_count * int(pcs_per_carton)


def compact_number(value):
    number = float(value or 0)
    if number.is_integer():
        return str(int(number))
    return (f"{number:.3f}").rstrip("0").rstrip(".")


def product_payload_to_values(data: dict):
    carton_length = to_float(data.get("carton_length"), 0.0)
    carton_width = to_float(data.get("carton_width"), 0.0)
    carton_height = to_float(data.get("carton_height"), 0.0)

    calculated_cbm = calculate_carton_cbm(
        carton_length,
        carton_width,
        carton_height,
    )

    # L/W/H are the source of truth when all three are available. The incoming
    # CBM is retained only for legacy / partially-filled records.
    carton_cbm = calculated_cbm or to_float(data.get("carton_cbm"), 0.0)

    pcs_per_carton = to_int(data.get("pcs_per_carton"), 0)

    def loading_value(key, container_key):
        manual_value = to_int(data.get(key), 0)
        if manual_value > 0:
            return manual_value
        return calculate_container_loading_pcs(
            carton_cbm,
            CONTAINER_USABLE_CBM[container_key],
            pcs_per_carton,
        )

    legacy_volume = str(data.get("volume", "") or "").strip()
    if carton_length > 0 and carton_width > 0 and carton_height > 0:
        legacy_volume = (
            f"{compact_number(carton_length)} x "
            f"{compact_number(carton_width)} x "
            f"{compact_number(carton_height)} cm"
        )

    return {
        "name": str(data.get("name", "") or "").strip(),
        "description": str(data.get("description", "") or "").strip(),
        "slogan": str(data.get("slogan", "") or "").strip(),
        "image": str(data.get("image", "") or "").strip(),
        "image_2": str(data.get("image_2", "") or "").strip(),
        "image_3": str(data.get("image_3", "") or "").strip(),
        "image_4": str(data.get("image_4", "") or "").strip(),
        "image_5": str(data.get("image_5", "") or "").strip(),
        "image_6": str(data.get("image_6", "") or "").strip(),
        "video": str(data.get("video", "") or "").strip(),
        "moq": to_int(data.get("moq"), 0),
        "material": str(data.get("material", "") or "").strip(),
        "volume": legacy_volume,
        "carton_length": carton_length,
        "carton_width": carton_width,
        "carton_height": carton_height,
        "carton_cbm": carton_cbm,
        "pcs_per_carton": pcs_per_carton,
        "loading_20gp": loading_value("loading_20gp", "20gp"),
        "loading_40gp": loading_value("loading_40gp", "40gp"),
        "loading_40hq": loading_value("loading_40hq", "40hq"),
        "nearest_port": str(data.get("nearest_port", "") or "").strip(),
        "lead_time": str(data.get("lead_time", "") or "").strip(),
        "size": str(data.get("size", "") or "").strip(),
        "packing": str(data.get("packing", "") or "").strip(),
        "category": str(data.get("category", "") or "").strip(),
        "price": to_float(data.get("price"), 0.0),
        "price_500": to_float(data.get("price_500"), 0.0),
        "price_1000": to_float(data.get("price_1000"), 0.0),
        "price_3000": to_float(data.get("price_3000"), 0.0),
        "price_10000": to_float(data.get("price_10000"), 0.0),
        "price_50000": to_float(data.get("price_50000"), 0.0),
    }


# =========================
# ADMIN EXCEL PRODUCT IMPORT
# =========================

EXCEL_IMPORT_MAX_BYTES = 100 * 1024 * 1024
EXCEL_IMPORT_MAX_ROWS = 2000

# Standard/simple template aliases remain supported as a fallback.
EXCEL_PRODUCT_FIELD_ALIASES = {
    "name": ["Product Name", "Name", "Item Name", "Product"],
    "category": ["Category", "Product Category"],
    "slogan": ["Product Slogan / Badge", "Product Slogan", "Slogan", "Badge"],
    "moq": ["MOQ", "Minimum Order Quantity", "Min Order Qty"],
    "material": ["Material", "Materials"],
    "size": ["Product Dimension", "Product Dimensions", "Product Size", "Size"],
    "packing": ["Packing", "Packaging", "Packing Method"],
    "volume": [
        "Carton Dimension", "Carton Dimensions", "Carton Size", "CTN Size",
        "Carton Dimension (cm)", "Carton Size (cm)",
    ],
    "carton_length": [
        "Carton L", "Carton L (cm)", "Carton Length", "Carton Length (cm)",
        "CTN L", "CTN Length",
    ],
    "carton_width": [
        "Carton W", "Carton W (cm)", "Carton Width", "Carton Width (cm)",
        "CTN W", "CTN Width",
    ],
    "carton_height": [
        "Carton H", "Carton H (cm)", "Carton Height", "Carton Height (cm)",
        "CTN H", "CTN Height",
    ],
    "carton_cbm": ["CBM", "CBM / Carton", "Carton CBM", "CBM/CTN", "CTN CBM"],
    "pcs_per_carton": [
        "PCS/CTN", "PCS / CTN", "Pcs/Ctn", "PCS per Carton",
        "Qty/Carton", "Qty / Carton", "Quantity per Carton", "Carton Qty",
    ],
    "loading_20gp": [
        "20GP", "20GP Loading Qty", "20GP Loading", "20FT", "20FT Qty",
        "20FC Loading", "20FC",
    ],
    "loading_40gp": [
        "40GP", "40GP Loading Qty", "40GP Loading", "40FT", "40FT Qty",
        "40FC Loading", "40FC",
    ],
    "loading_40hq": [
        "40HQ", "40HQ Loading Qty", "40HQ Loading", "40HC", "40HC Qty",
        "40HC Loading",
    ],
    "nearest_port": ["Nearest Port", "Port", "Loading Port", "Port of Loading"],
    "lead_time": ["Lead Time", "Leadtime", "Production Lead Time", "Delivery Time"],
    "price": ["Base Price", "Price", "Unit Price", "FOB Price", "FOB"],
    "price_500": ["Price 500+", "Price 500", "500+ Price", "500 Price"],
    "price_1000": ["Price 1000+", "Price 1000", "1000+ Price", "1000 Price"],
    "price_3000": ["Price 3000+", "Price 3000", "3000+ Price", "3000 Price"],
    "price_10000": ["Price 10000+", "Price 10000", "10000+ Price", "10000 Price"],
    "price_50000": ["Price 50000+", "Price 50000", "50000+ Price", "50000 Price"],
    "description": ["Specification", "Specifications", "Spec", "Product Description"],
}

EXCEL_IMPORT_TEMPLATE_HEADERS = [
    "ITEM NUMBER", "UPC CODE", "BRAND", "IMAGE", "PACKAGING DEMO",
    "DESCRIPTION", "MATERIAL", "SPECIFICATION", "PACKAGING METHOD",
    "PROMOTION QTY", "PRICE (USD)", "STANDARD QTY", "PRICE (USD)",
    "INNER CARTON", "", "", "",
    "OUTER CARTON", "", "", "",
    "CBM", "20FC LOADING", "40FC LOADING", "40HC LOADING",
    "PORT", "CONFIRMATION DDL", "LEADTIME", "REMARK", "REMARK",
]


def normalize_excel_header(value):
    text_value = str(value or "").strip().lower()
    return re.sub(r"[^a-z0-9]+", "", text_value)


def build_excel_header_lookup():
    lookup = {}
    for website_field, aliases in EXCEL_PRODUCT_FIELD_ALIASES.items():
        for alias in aliases:
            normalized = normalize_excel_header(alias)
            if normalized:
                lookup[normalized] = website_field
    return lookup


EXCEL_HEADER_LOOKUP = build_excel_header_lookup()


def excel_json_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def normalize_excel_numeric_value(value):
    if value in (None, ""):
        return ""
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return value

    raw = str(value).strip().replace(",", "")
    if raw.upper() in {"N/A", "NA", "TBD", "/", "-", "--"}:
        return ""

    match = re.search(r"-?(?:\d+(?:\.\d+)?|\.\d+)", raw)
    if not match:
        return ""

    try:
        number = float(match.group(0))
        return int(number) if number.is_integer() else number
    except Exception:
        return ""


def parse_carton_dimension_text(value):
    raw = str(value or "").strip()
    if not raw:
        return None
    numbers = re.findall(r"\d+(?:\.\d+)?", raw.replace(",", "."))
    if len(numbers) < 3:
        return None
    try:
        return tuple(float(number) for number in numbers[:3])
    except Exception:
        return None


def append_multiline(existing, addition):
    existing_text = str(existing or "").strip()
    addition_text = str(addition or "").strip()
    if not addition_text:
        return existing_text
    if not existing_text:
        return addition_text
    return existing_text + "\n" + addition_text


async def save_excel_request_to_temp(request: Request):
    temp_file = tempfile.NamedTemporaryFile(mode="wb", suffix=".xlsx", delete=False)
    temp_path = Path(temp_file.name)
    total_bytes = 0

    try:
        async for chunk in request.stream():
            if not chunk:
                continue
            total_bytes += len(chunk)
            if total_bytes > EXCEL_IMPORT_MAX_BYTES:
                raise ValueError("Excel file is too large. Maximum size is 100MB.")
            temp_file.write(chunk)

        temp_file.flush()
        temp_file.close()

        if total_bytes <= 0:
            raise ValueError("Excel file is empty.")

        return temp_path, total_bytes
    except Exception:
        try:
            temp_file.close()
        except Exception:
            pass
        try:
            temp_path.unlink(missing_ok=True)
        except Exception:
            pass
        raise


def load_excel_workbook(file_path, data_only=True):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel import requires openpyxl. Add openpyxl==3.1.5 to requirements.txt and redeploy."
        ) from exc

    try:
        return load_workbook(
            filename=str(file_path),
            read_only=True,
            data_only=data_only,
        )
    except Exception as exc:
        raise ValueError("The uploaded file is not a valid .xlsx workbook.") from exc


def find_standard_excel_header_row(worksheet):
    best = None
    max_scan_rows = min(int(worksheet.max_row or 0), 25)

    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
        start=1,
    ):
        mapping = {}
        recognized = 0

        for column_index, value in enumerate(values, start=1):
            normalized = normalize_excel_header(value)
            website_field = EXCEL_HEADER_LOOKUP.get(normalized)
            if website_field and website_field not in mapping.values():
                mapping[column_index] = website_field
                recognized += 1

        has_name = "name" in mapping.values()
        score = recognized + (100 if has_name else 0)

        if best is None or score > best["score"]:
            best = {
                "row_number": row_number,
                "mapping": mapping,
                "values": list(values),
                "recognized": recognized,
                "has_name": has_name,
                "score": score,
            }

    if not best or not best["has_name"]:
        raise ValueError("Could not find a Product Name column.")
    return best


def find_copec_header_row(worksheet):
    max_scan_rows = min(int(worksheet.max_row or 0), 20)

    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
        start=1,
    ):
        normalized_values = [normalize_excel_header(value) for value in values]
        required = {
            "itemnumber", "description", "material", "specification",
            "packagingmethod", "outercarton", "cbm", "port", "leadtime",
        }

        if not required.issubset(set(normalized_values)):
            continue

        positions = {}
        for column_index, normalized in enumerate(normalized_values, start=1):
            if normalized and normalized not in positions:
                positions[normalized] = column_index

        price_columns = [
            index for index, normalized in enumerate(normalized_values, start=1)
            if normalized in {"priceusd", "price"}
        ]
        if not price_columns:
            continue

        return {
            "row_number": row_number,
            "values": list(values),
            "positions": positions,
            "price_columns": price_columns[:2],
        }

    return None


def excel_cell(values, column_index):
    if not column_index or column_index <= 0:
        return ""
    index = column_index - 1
    if index >= len(values):
        return ""
    return excel_json_value(values[index])


def copec_price_values(values, header_info):
    pairs = []

    for price_column in header_info.get("price_columns", []):
        quantity_column = price_column - 1
        quantity = normalize_excel_numeric_value(excel_cell(values, quantity_column))
        price = normalize_excel_numeric_value(excel_cell(values, price_column))
        if price not in ("", None):
            pairs.append({"qty": quantity, "price": price})

    payload = {
        "moq": "",
        "price": "",
        "price_500": "",
        "price_1000": "",
        "price_3000": "",
        "price_10000": "",
        "price_50000": "",
    }

    quantity_pairs = [
        pair for pair in pairs
        if normalize_excel_numeric_value(pair.get("qty")) not in ("", None)
        and float(normalize_excel_numeric_value(pair.get("qty"))) > 0
    ]

    if quantity_pairs:
        quantity_pairs.sort(
            key=lambda pair: float(normalize_excel_numeric_value(pair.get("qty")))
        )
        base_pair = quantity_pairs[0]
        payload["moq"] = normalize_excel_numeric_value(base_pair.get("qty"))
        payload["price"] = normalize_excel_numeric_value(base_pair.get("price"))

        tier_map = {
            500: "price_500",
            1000: "price_1000",
            3000: "price_3000",
            10000: "price_10000",
            50000: "price_50000",
        }

        for pair in quantity_pairs:
            qty_value = normalize_excel_numeric_value(pair.get("qty"))
            price_value = normalize_excel_numeric_value(pair.get("price"))
            try:
                qty_int = int(float(qty_value))
            except Exception:
                continue
            tier_key = tier_map.get(qty_int)
            if tier_key and price_value not in ("", None):
                payload[tier_key] = price_value
    elif pairs:
        payload["price"] = normalize_excel_numeric_value(pairs[0].get("price"))

    return payload, pairs


def copec_row_has_anchor(values, header_info):
    positions = header_info["positions"]
    item_number = str(excel_cell(values, positions.get("itemnumber"))).strip()

    _, raw_pairs = copec_price_values(values, header_info)
    has_price = any(
        normalize_excel_numeric_value(pair.get("price")) not in ("", None)
        for pair in raw_pairs
    )

    outer_start = positions.get("outercarton")
    has_outer_carton = False
    if outer_start:
        for column_index in range(outer_start, outer_start + 4):
            if normalize_excel_numeric_value(excel_cell(values, column_index)) not in ("", None):
                has_outer_carton = True
                break

    return bool(item_number or has_price or has_outer_carton)


def copec_row_to_payload(values, header_info, sheet_name):
    positions = header_info["positions"]

    description_text = excel_cell(values, positions.get("description"))
    material_text = excel_cell(values, positions.get("material"))
    specification_text = excel_cell(values, positions.get("specification"))
    packing_text = excel_cell(values, positions.get("packagingmethod"))

    outer_start = positions.get("outercarton")
    outer_qty = normalize_excel_numeric_value(excel_cell(values, outer_start)) if outer_start else ""
    carton_length = normalize_excel_numeric_value(excel_cell(values, outer_start + 1)) if outer_start else ""
    carton_width = normalize_excel_numeric_value(excel_cell(values, outer_start + 2)) if outer_start else ""
    carton_height = normalize_excel_numeric_value(excel_cell(values, outer_start + 3)) if outer_start else ""

    price_payload, _ = copec_price_values(values, header_info)

    payload = {
        "name": str(description_text or "").strip(),
        "category": str(sheet_name or "").strip(),
        "slogan": "",
        "material": str(material_text or "").strip(),
        "size": "",
        "packing": str(packing_text or "").strip(),
        "volume": "",
        "carton_length": carton_length,
        "carton_width": carton_width,
        "carton_height": carton_height,
        "carton_cbm": normalize_excel_numeric_value(excel_cell(values, positions.get("cbm"))),
        "pcs_per_carton": outer_qty,
        "loading_20gp": normalize_excel_numeric_value(excel_cell(values, positions.get("20fcloading"))),
        "loading_40gp": normalize_excel_numeric_value(excel_cell(values, positions.get("40fcloading"))),
        "loading_40hq": normalize_excel_numeric_value(excel_cell(values, positions.get("40hcloading"))),
        "nearest_port": str(excel_cell(values, positions.get("port")) or "").strip(),
        "lead_time": str(excel_cell(values, positions.get("leadtime")) or "").strip(),
        "description": str(specification_text or "").strip(),
        **price_payload,
    }

    if all([
        normalize_excel_numeric_value(carton_length) not in ("", None),
        normalize_excel_numeric_value(carton_width) not in ("", None),
        normalize_excel_numeric_value(carton_height) not in ("", None),
    ]):
        payload["volume"] = f"{carton_length} x {carton_width} x {carton_height} cm"

    return payload


def merge_copec_continuation(current_record, values, header_info, excel_row_number):
    if current_record is None:
        return None

    positions = header_info["positions"]
    payload = current_record["data"]

    component_name = str(excel_cell(values, positions.get("description")) or "").strip()
    component_material = str(excel_cell(values, positions.get("material")) or "").strip()
    component_spec = str(excel_cell(values, positions.get("specification")) or "").strip()
    component_packing = str(excel_cell(values, positions.get("packagingmethod")) or "").strip()

    payload["name"] = append_multiline(payload.get("name"), component_name)
    payload["material"] = append_multiline(payload.get("material"), component_material)

    component_detail = f"{component_name}: {component_spec}" if component_name and component_spec else component_spec
    payload["description"] = append_multiline(payload.get("description"), component_detail)

    if not str(payload.get("packing", "") or "").strip() and component_packing:
        payload["packing"] = component_packing

    current_record["excel_row_end"] = excel_row_number
    return current_record


def parse_copec_sheet(worksheet, header_info, remaining_limit):
    header_row_number = header_info["row_number"]
    data_start_row = header_row_number + 2
    records = []
    current_record = None

    inherited_keys = [
        "name", "material", "description", "packing", "size", "volume",
        "carton_length", "carton_width", "carton_height", "carton_cbm",
        "pcs_per_carton", "loading_20gp", "loading_40gp", "loading_40hq",
        "nearest_port", "lead_time",
    ]

    for excel_row_number, values in enumerate(
        worksheet.iter_rows(min_row=data_start_row, values_only=True),
        start=data_start_row,
    ):
        description_text = str(excel_cell(values, header_info["positions"].get("description")) or "").strip()
        material_text = str(excel_cell(values, header_info["positions"].get("material")) or "").strip()
        specification_text = str(excel_cell(values, header_info["positions"].get("specification")) or "").strip()

        if not any([description_text, material_text, specification_text]) and not copec_row_has_anchor(values, header_info):
            continue

        anchored = copec_row_has_anchor(values, header_info)

        if not anchored and current_record is not None:
            merge_copec_continuation(current_record, values, header_info, excel_row_number)
            continue

        payload = copec_row_to_payload(values, header_info, worksheet.title)

        # COPEC sheets frequently leave DESCRIPTION / SPECIFICATION / carton data
        # blank on material or price variants because those cells visually belong
        # to the product above. Inherit those shared values while keeping the
        # variant's own material / MOQ / price.
        if current_record is not None:
            previous_payload = current_record.get("data", {})
            for key in inherited_keys:
                if payload.get(key) in ("", None, 0, 0.0):
                    previous_value = previous_payload.get(key)
                    if previous_value not in ("", None, 0, 0.0):
                        payload[key] = previous_value

        row_errors = []
        if not str(payload.get("name", "") or "").strip():
            row_errors.append("DESCRIPTION is required")

        record = {
            "sheet_name": worksheet.title,
            "excel_row": excel_row_number,
            "excel_row_end": excel_row_number,
            "data": payload,
            "errors": row_errors,
            "valid": not row_errors,
        }
        records.append(record)
        current_record = record

        if len(records) >= remaining_limit:
            break

    return records


def parse_standard_sheet(worksheet, header_info, remaining_limit):
    header_row_number = header_info["row_number"]
    column_mapping = header_info["mapping"]
    records = []

    for excel_row_number, values in enumerate(
        worksheet.iter_rows(min_row=header_row_number + 1, values_only=True),
        start=header_row_number + 1,
    ):
        payload = {}
        any_value = False

        for column_index, website_field in column_mapping.items():
            raw_value = values[column_index - 1] if column_index - 1 < len(values) else None
            value = excel_json_value(raw_value)
            payload[website_field] = value
            if value not in ("", None):
                any_value = True

        if not any_value:
            continue

        for numeric_key in [
            "moq", "carton_length", "carton_width", "carton_height",
            "carton_cbm", "pcs_per_carton", "loading_20gp",
            "loading_40gp", "loading_40hq", "price", "price_500",
            "price_1000", "price_3000", "price_10000", "price_50000",
        ]:
            if numeric_key in payload:
                payload[numeric_key] = normalize_excel_numeric_value(payload.get(numeric_key))

        if not all([
            normalize_excel_numeric_value(payload.get("carton_length")),
            normalize_excel_numeric_value(payload.get("carton_width")),
            normalize_excel_numeric_value(payload.get("carton_height")),
        ]):
            combined_dimensions = parse_carton_dimension_text(payload.get("volume"))
            if combined_dimensions:
                payload["carton_length"], payload["carton_width"], payload["carton_height"] = combined_dimensions

        row_errors = []
        if not str(payload.get("name", "") or "").strip():
            row_errors.append("Product Name is required")

        records.append({
            "sheet_name": worksheet.title,
            "excel_row": excel_row_number,
            "excel_row_end": excel_row_number,
            "data": payload,
            "errors": row_errors,
            "valid": not row_errors,
        })

        if len(records) >= remaining_limit:
            break

    return records


def copec_matched_columns():
    return [
        {"excel_column": "Sheet Name", "website_field": "Category"},
        {"excel_column": "DESCRIPTION", "website_field": "Product Name"},
        {"excel_column": "MATERIAL", "website_field": "Material"},
        {"excel_column": "SPECIFICATION", "website_field": "Specification"},
        {"excel_column": "PACKAGING METHOD", "website_field": "Packing"},
        {"excel_column": "QTY + PRICE (USD)", "website_field": "MOQ / Base Price / Price Tiers"},
        {"excel_column": "OUTER CARTON QTY", "website_field": "PCS/CTN"},
        {"excel_column": "OUTER CARTON L/W/H", "website_field": "Carton L/W/H"},
        {"excel_column": "CBM", "website_field": "CBM / Carton"},
        {"excel_column": "20FC LOADING", "website_field": "20GP Loading Qty (PCS)"},
        {"excel_column": "40FC LOADING", "website_field": "40GP Loading Qty (PCS)"},
        {"excel_column": "40HC LOADING", "website_field": "40HQ Loading Qty (PCS)"},
        {"excel_column": "PORT", "website_field": "Nearest Port"},
        {"excel_column": "LEADTIME", "website_field": "Lead Time"},
    ]


def parse_product_excel(file_path):
    workbook = load_excel_workbook(file_path, data_only=True)

    try:
        all_records = []
        sheet_summaries = []
        copec_detected = False

        for worksheet in workbook.worksheets:
            header_info = find_copec_header_row(worksheet)
            if not header_info:
                continue

            copec_detected = True
            remaining = EXCEL_IMPORT_MAX_ROWS - len(all_records)
            if remaining <= 0:
                break

            records = parse_copec_sheet(worksheet, header_info, remaining)
            if records:
                all_records.extend(records)
                sheet_summaries.append({
                    "sheet_name": worksheet.title,
                    "header_row": header_info["row_number"],
                    "products": len(records),
                    "category": worksheet.title,
                })

        matched_columns = []
        unmatched_columns = []
        import_format = "COPEC multi-sheet format"

        if not copec_detected:
            import_format = "Standard product import format"
            best_sheet = None
            best_header = None

            for worksheet in workbook.worksheets:
                try:
                    header_info = find_standard_excel_header_row(worksheet)
                except Exception:
                    continue

                if best_header is None or header_info.get("recognized", 0) > best_header.get("recognized", 0):
                    best_sheet = worksheet
                    best_header = header_info

            if best_sheet is None or best_header is None:
                raise ValueError(
                    "Could not recognize the workbook format. Use the COPEC quotation layout with "
                    "DESCRIPTION / MATERIAL / SPECIFICATION / OUTER CARTON / CBM / PORT / LEADTIME columns."
                )

            all_records = parse_standard_sheet(best_sheet, best_header, EXCEL_IMPORT_MAX_ROWS)
            sheet_summaries = [{
                "sheet_name": best_sheet.title,
                "header_row": best_header["row_number"],
                "products": len(all_records),
                "category": "",
            }]

            for column_index, header in enumerate(best_header["values"], start=1):
                header_text = str(header or "").strip()
                if not header_text:
                    continue
                website_field = best_header["mapping"].get(column_index)
                if website_field:
                    matched_columns.append({
                        "excel_column": header_text,
                        "website_field": website_field,
                    })
                else:
                    unmatched_columns.append(header_text)
        else:
            matched_columns = copec_matched_columns()
            unmatched_columns = [
                "ITEM NUMBER", "UPC CODE", "BRAND", "IMAGE",
                "PACKAGING DEMO", "INNER CARTON", "CONFIRMATION DDL", "REMARK",
            ]

        if not all_records:
            raise ValueError("No product rows were found in the workbook.")

        valid_rows = [record["data"] for record in all_records if record.get("valid")]
        error_records = [record for record in all_records if not record.get("valid")]

        return {
            "import_format": import_format,
            "sheet_name": ", ".join(summary["sheet_name"] for summary in sheet_summaries),
            "sheet_names": [summary["sheet_name"] for summary in sheet_summaries],
            "sheet_summaries": sheet_summaries,
            "header_row": sheet_summaries[0]["header_row"] if sheet_summaries else 1,
            "matched_columns": matched_columns,
            "unmatched_columns": unmatched_columns,
            "rows": valid_rows,
            "preview_rows": all_records[:30],
            "valid_rows": len(valid_rows),
            "error_rows": len(error_records),
            "total_rows": len(all_records),
            "truncated": len(all_records) >= EXCEL_IMPORT_MAX_ROWS,
        }
    finally:
        workbook.close()


@app.post("/admin/products/import-excel/preview")
async def preview_product_excel(
    request: Request,
    requester_username: str = "",
    filename: str = "",
):
    db = SessionLocal()
    temp_path = None

    try:
        requester_username = str(requester_username or "").strip()
        if not can_approved_admin(db, requester_username):
            return {
                "success": False,
                "message": "Only an approved admin can import products from Excel."
            }

        filename = str(filename or "").strip()
        if not filename.lower().endswith(".xlsx"):
            return {
                "success": False,
                "message": "Please upload an .xlsx Excel file."
            }

        temp_path, file_size = await save_excel_request_to_temp(request)
        parsed = parse_product_excel(temp_path)

        return {
            "success": True,
            "filename": filename,
            "file_size": file_size,
            **parsed,
        }
    except Exception as exc:
        return {
            "success": False,
            "message": str(exc),
        }
    finally:
        db.close()
        if temp_path:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except Exception:
                pass


@app.post("/admin/products/import-excel/confirm")
def confirm_product_excel_import(data: dict):
    db = SessionLocal()

    try:
        requester_username = str(data.get("requester_username", "") or "").strip()
        if not can_approved_admin(db, requester_username):
            return {
                "success": False,
                "message": "Only an approved admin can import products from Excel."
            }

        rows = data.get("rows", [])
        if not isinstance(rows, list) or not rows:
            return {
                "success": False,
                "message": "There are no valid Excel rows to import."
            }

        if len(rows) > EXCEL_IMPORT_MAX_ROWS:
            return {
                "success": False,
                "message": f"A maximum of {EXCEL_IMPORT_MAX_ROWS} products can be imported at one time."
            }

        next_id = db.execute(
            text("SELECT COALESCE(MAX(id), 0) + 1 FROM products")
        ).scalar()
        next_id = int(next_id or 1)

        created = []
        skipped = []

        for index, raw_row in enumerate(rows, start=1):
            if not isinstance(raw_row, dict):
                skipped.append({"row": index, "reason": "Invalid row format"})
                continue

            values = product_payload_to_values(raw_row)
            if not values["name"]:
                skipped.append({"row": index, "reason": "Product Name is required"})
                continue

            product = Product(id=next_id, **values)
            product.is_active = 1
            db.add(product)
            created.append({"id": next_id, "name": values["name"]})
            next_id += 1

        if not created:
            db.rollback()
            return {
                "success": False,
                "message": "No valid products were imported.",
                "skipped": skipped,
            }

        db.commit()

        return {
            "success": True,
            "message": f"Imported {len(created)} product(s) from Excel.",
            "created_count": len(created),
            "skipped_count": len(skipped),
            "created": created[:50],
            "skipped": skipped[:50],
        }

    except Exception as exc:
        db.rollback()
        return {
            "success": False,
            "message": "Excel import failed: " + str(exc),
        }
    finally:
        db.close()


@app.get("/admin/products/import-excel/template")
def download_product_excel_template():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return {
            "success": False,
            "message": "Excel template requires openpyxl. Add openpyxl==3.1.5 to requirements.txt and redeploy."
        }

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CATEGORY NAME"

    worksheet["F1"] = "BILL C.TONG"
    worksheet["H1"] = "BILL.TONG@LINK-INT.COM"
    worksheet["N1"] = "DATE OF QUOTATION"
    worksheet["N2"] = "DATE OF EXPIRATION"
    worksheet["N3"] = "INCOTERM OF TRADE"
    worksheet["V3"] = "FOB"

    for column_index, value in enumerate(EXCEL_IMPORT_TEMPLATE_HEADERS, start=1):
        worksheet.cell(row=4, column=column_index, value=value)

    worksheet.merge_cells("N4:Q4")
    worksheet["N4"] = "INNER CARTON"
    worksheet.merge_cells("R4:U4")
    worksheet["R4"] = "OUTER CARTON"

    for column_index, value in enumerate(["QTY", "L", "W", "H"], start=14):
        worksheet.cell(row=5, column=column_index, value=value)
    for column_index, value in enumerate(["QTY", "L", "W", "H"], start=18):
        worksheet.cell(row=5, column=column_index, value=value)

    sample_values = [
        "FU026-001", "TBD", "TBD", "", "",
        "MICROFIBER CLEANING CLOTH",
        "80% POLYESTER\n20% POLYAMIDE",
        "PRODUCT SIZE: 30*30CM",
        "COLOUR SLEEVE",
        50000, 0.55, 3000, 0.62,
        "/", "/", "/", "/",
        72, 52, 41, 36,
        0.076752,
        26208, 56448, 66528,
        "NINGBO", "N/A", "30-35 DAYS", "", "",
    ]

    for column_index, value in enumerate(sample_values, start=1):
        worksheet.cell(row=6, column=column_index, value=value)

    header_fill = PatternFill("solid", fgColor="1F3C88")
    header_font = Font(color="FFFFFF", bold=True)

    for row_number in [4, 5]:
        for cell in worksheet[row_number]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    for cell in worksheet[6]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A6"
    worksheet.row_dimensions[4].height = 42
    worksheet.row_dimensions[5].height = 24
    worksheet.row_dimensions[6].height = 70

    widths = {
        "A": 18, "B": 14, "C": 14, "D": 16, "E": 18,
        "F": 32, "G": 28, "H": 34, "I": 24,
        "J": 15, "K": 14, "L": 15, "M": 14,
        "N": 10, "O": 10, "P": 10, "Q": 10,
        "R": 10, "S": 10, "T": 10, "U": 10,
        "V": 12, "W": 16, "X": 16, "Y": 16,
        "Z": 16, "AA": 18, "AB": 18, "AC": 24, "AD": 24,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=COPEC_Product_Import_Template.xlsx"
        },
    )


@app.get("/admin/products")
def get_admin_products():
    db = SessionLocal()

    products = db.query(Product).order_by(Product.id.desc()).all()

    result = [
        serialize_product(product, 1.0)
        for product in products
    ]

    db.close()
    return result


@app.post("/admin/products")
def create_admin_product(data: dict):
    db = SessionLocal()

    try:
        values = product_payload_to_values(data)

        if not values["name"]:
            db.close()
            return {
                "success": False,
                "message": "Product name is required."
            }

        # Products imported from Excel may already have fixed IDs.
        # Creating new products without syncing the DB sequence can cause
        # duplicate-key errors on PostgreSQL, so we assign the next ID safely.
        next_id = db.execute(
            text("SELECT COALESCE(MAX(id), 0) + 1 FROM products")
        ).scalar()

        product = Product(id=int(next_id or 1), **values)
        product.is_active = 1

        db.add(product)
        db.commit()

        product_id = product.id
        db.close()

        return {
            "success": True,
            "message": f"Product #{product_id} has been created."
        }

    except Exception as e:
        db.rollback()
        db.close()
        return {
            "success": False,
            "message": "Create product failed: " + str(e)
        }


@app.post("/admin/products/{product_id}")
def update_admin_product(product_id: int, data: dict):
    db = SessionLocal()

    product = db.query(Product).filter(
        Product.id == product_id
    ).first()

    if not product:
        db.close()
        return {
            "success": False,
            "message": "Product not found."
        }

    values = product_payload_to_values(data)

    if not values["name"]:
        db.close()
        return {
            "success": False,
            "message": "Product name cannot be empty."
        }

    for key, value in values.items():
        setattr(product, key, value)

    db.commit()
    db.close()

    return {
        "success": True,
        "message": f"Product #{product_id} has been updated."
    }


@app.post("/admin/products/{product_id}/toggle")
def toggle_admin_product(product_id: int, data: dict):
    db = SessionLocal()

    product = db.query(Product).filter(
        Product.id == product_id
    ).first()

    if not product:
        db.close()
        return {
            "success": False,
            "message": "Product not found."
        }

    active_value = data.get("is_active", 1)

    if active_value in [False, "false", "False", "0", 0]:
        product.is_active = 0
        message = f"Product #{product_id} has been hidden."
    else:
        product.is_active = 1
        message = f"Product #{product_id} has been activated."

    db.commit()
    db.close()

    return {
        "success": True,
        "message": message
    }


@app.post("/admin/products/bulk-delete")
def bulk_delete_admin_products(data: dict):
    raw_ids = data.get("product_ids", [])

    if not isinstance(raw_ids, list):
        return {
            "success": False,
            "message": "product_ids must be a list."
        }

    product_ids = []
    seen_ids = set()

    for raw_id in raw_ids:
        try:
            product_id = int(raw_id)
        except (TypeError, ValueError):
            continue

        if product_id <= 0 or product_id in seen_ids:
            continue

        seen_ids.add(product_id)
        product_ids.append(product_id)

    if not product_ids:
        return {
            "success": False,
            "message": "Please select at least one valid product."
        }

    if len(product_ids) > 2000:
        return {
            "success": False,
            "message": "A maximum of 2,000 products can be deleted at one time."
        }

    db = SessionLocal()

    try:
        products = db.query(Product).filter(
            Product.id.in_(product_ids)
        ).all()

        found_ids = {int(product.id) for product in products}
        missing_ids = [product_id for product_id in product_ids if product_id not in found_ids]

        for product in products:
            db.delete(product)

        db.commit()

        deleted_count = len(products)

        return {
            "success": True,
            "deleted_count": deleted_count,
            "missing_ids": missing_ids,
            "message": f"Deleted {deleted_count} product{'s' if deleted_count != 1 else ''}."
        }
    except Exception as exc:
        db.rollback()
        print("Bulk product delete failed:", exc)
        return {
            "success": False,
            "message": "Bulk product delete failed."
        }
    finally:
        db.close()


@app.post("/admin/products/{product_id}/delete")
def delete_admin_product(product_id: int):
    db = SessionLocal()

    product = db.query(Product).filter(
        Product.id == product_id
    ).first()

    if not product:
        db.close()
        return {
            "success": False,
            "message": "Product not found."
        }

    db.delete(product)
    db.commit()
    db.close()

    return {
        "success": True,
        "message": f"Product #{product_id} has been deleted."
    }


def safe_upload_extension(filename, media_type):
    suffix = Path(str(filename or "")).suffix.lower()

    image_extensions = [".jpg", ".jpeg", ".png", ".webp", ".gif"]
    video_extensions = [".mp4", ".webm", ".mov", ".m4v"]

    if media_type == "video":
        if suffix not in video_extensions:
            return None
    else:
        if suffix not in image_extensions:
            return None

    return suffix


def safe_shared_file_extension(filename):
    suffix = Path(str(filename or "")).suffix.lower()
    allowed_extensions = {
        ".jpg", ".jpeg", ".png", ".webp", ".gif",
        ".pdf", ".doc", ".docx", ".xls", ".xlsx",
        ".ppt", ".pptx", ".csv", ".txt", ".zip", ".rar",
        ".mp4", ".webm", ".mov", ".m4v"
    }
    return suffix if suffix in allowed_extensions else None


def get_shared_file_type(filename):
    suffix = Path(str(filename or "")).suffix.lower()

    if suffix in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        return "Image"
    if suffix in {".mp4", ".webm", ".mov", ".m4v"}:
        return "Video"
    if suffix == ".pdf":
        return "PDF"
    if suffix in {".doc", ".docx"}:
        return "Word"
    if suffix in {".xls", ".xlsx", ".csv"}:
        return "Excel"
    if suffix in {".ppt", ".pptx"}:
        return "PowerPoint"
    if suffix in {".zip", ".rar"}:
        return "Archive"
    return "Document"


def serialize_shared_file(shared_file):
    return {
        "id": shared_file.id,
        "folder_id": int(getattr(shared_file, "folder_id", 0) or 0),
        "original_name": shared_file.original_name or "",
        "file_path": shared_file.file_path or "",
        "file_url": shared_file.file_path or "",
        "file_type": shared_file.file_type or "Document",
        "file_size": int(shared_file.file_size or 0),
        "uploaded_by": shared_file.uploaded_by or "",
        "uploaded_at": shared_file.uploaded_at or ""
    }


def can_access_shared_files(db, username, admin_only=False):
    user = db.query(User).filter(
        User.username == str(username or "").strip()
    ).first()

    if not user:
        return False

    if user.role == "admin":
        return True

    return not admin_only and user.role == "sales"


def can_manage_shared_files(db, username):
    normalized_username = str(username or "").strip().lower()
    user = get_user_by_username_case_insensitive(db, normalized_username)

    if not user:
        return False

    role = str(user.role or "").strip().lower()
    approval_status = str(
        user.approval_status or "Approved"
    ).strip()

    if approval_status != "Approved":
        return False

    if role == "admin":
        return True

    return normalized_username == "bill" and role == "sales"


def get_shared_sales_folders(db):
    count_rows = db.query(
        SharedSalesFile.folder_id,
        func.count(SharedSalesFile.id)
    ).group_by(SharedSalesFile.folder_id).all()

    counts = {
        int(folder_id or 0): int(file_count or 0)
        for folder_id, file_count in count_rows
    }

    folders = db.query(SharedSalesFolder).order_by(
        SharedSalesFolder.name.asc()
    ).all()

    result = [
        {
            "id": folder.id,
            "name": folder.name,
            "created_by": folder.created_by or "",
            "created_at": folder.created_at or "",
            "file_count": counts.pop(folder.id, 0)
        }
        for folder in folders
    ]

    unfiled_count = sum(counts.values())
    if unfiled_count:
        result.append({
            "id": 0,
            "name": "Unfiled Files",
            "created_by": "",
            "created_at": "",
            "file_count": unfiled_count,
            "is_unfiled": True
        })

    return result


@app.post("/admin/product-media-upload-chunk")
def upload_product_media_chunk(data: dict):
    upload_id = str(data.get("upload_id", "") or "").strip()
    filename = str(data.get("filename", "") or "").strip()
    media_type = str(data.get("media_type", "image") or "image").strip().lower()

    if media_type not in ["image", "video"]:
        media_type = "image"

    try:
        chunk_index = int(data.get("chunk_index", 0))
        total_chunks = int(data.get("total_chunks", 1))
    except Exception:
        return {
            "success": False,
            "message": "Invalid chunk information."
        }

    if not upload_id:
        return {
            "success": False,
            "message": "Missing upload id."
        }

    if chunk_index < 0 or total_chunks <= 0 or chunk_index >= total_chunks:
        return {
            "success": False,
            "message": "Invalid chunk index."
        }

    extension = safe_upload_extension(filename, media_type)

    if not extension:
        return {
            "success": False,
            "message": "Invalid file type."
        }

    chunk_data = str(data.get("chunk_data", "") or "")

    if "," in chunk_data:
        chunk_data = chunk_data.split(",", 1)[1]

    if not chunk_data:
        return {
            "success": False,
            "message": "Missing chunk data."
        }

    try:
        chunk_bytes = base64.b64decode(chunk_data)
    except Exception:
        return {
            "success": False,
            "message": "Invalid chunk data."
        }

    safe_upload_id = re.sub(r"[^a-zA-Z0-9_-]", "", upload_id)[:80]

    if not safe_upload_id:
        return {
            "success": False,
            "message": "Invalid upload id."
        }

    root_dir = Path("static") / "uploads" / "products"
    tmp_dir = root_dir / "_chunks" / safe_upload_id

    if media_type == "video":
        final_dir = root_dir / "videos"
    else:
        final_dir = root_dir / "images"

    try:
        tmp_dir.mkdir(parents=True, exist_ok=True)
        final_dir.mkdir(parents=True, exist_ok=True)

        chunk_path = tmp_dir / f"{chunk_index:06d}.part"
        chunk_path.write_bytes(chunk_bytes)

        if chunk_index < total_chunks - 1:
            return {
                "success": True,
                "message": f"Chunk {chunk_index + 1}/{total_chunks} uploaded.",
                "complete": False
            }

        missing_chunks = []

        for index in range(total_chunks):
            part_path = tmp_dir / f"{index:06d}.part"
            if not part_path.exists():
                missing_chunks.append(index)

        if missing_chunks:
            return {
                "success": False,
                "message": "Missing upload chunks: " + ",".join(str(i) for i in missing_chunks[:10])
            }

        final_name = (
            datetime.now().strftime("%Y%m%d%H%M%S") +
            "_" +
            uuid.uuid4().hex[:10] +
            extension
        )

        final_path = final_dir / final_name

        with final_path.open("wb") as output_file:
            for index in range(total_chunks):
                part_path = tmp_dir / f"{index:06d}.part"
                output_file.write(part_path.read_bytes())

        shutil.rmtree(tmp_dir, ignore_errors=True)

        public_url = "/" + str(final_path).replace("\\", "/")

        return {
            "success": True,
            "message": "File uploaded successfully.",
            "complete": True,
            "url": public_url
        }

    except Exception as e:
        return {
            "success": False,
            "message": "Upload failed: " + str(e)
        }


# =========================
# SHARED SALES FILE LIBRARY
# =========================

@app.post("/shared-sales-folders/create")
def create_shared_sales_folder(data: dict):
    db = SessionLocal()
    username = str(data.get("username", "") or "").strip()
    name = re.sub(r"\s+", " ", str(data.get("name", "") or "").strip())

    if not can_manage_shared_files(db, username):
        db.close()
        return {"success": False, "message": "Only BILL and approved admin users can create folders."}

    if not name:
        db.close()
        return {"success": False, "message": "Please enter a folder name."}

    if len(name) > 80:
        db.close()
        return {"success": False, "message": "Folder name must be 80 characters or less."}

    existing = db.query(SharedSalesFolder).filter(
        func.lower(SharedSalesFolder.name) == name.lower()
    ).first()

    if existing:
        db.close()
        return {"success": False, "message": "A folder with this name already exists."}

    folder = SharedSalesFolder(
        name=name,
        created_by=username,
        created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )
    db.add(folder)
    db.commit()

    result = {
        "success": True,
        "message": "Folder created.",
        "folder": {
            "id": folder.id,
            "name": folder.name,
            "created_by": folder.created_by,
            "created_at": folder.created_at,
            "file_count": 0
        }
    }
    db.close()
    return result


@app.post("/shared-sales-folders/{folder_id}/rename")
def rename_shared_sales_folder(folder_id: int, data: dict):
    db = SessionLocal()
    username = str(data.get("username", "") or "").strip()
    name = re.sub(r"\s+", " ", str(data.get("name", "") or "").strip())

    if not can_manage_shared_files(db, username):
        db.close()
        return {"success": False, "message": "Only BILL and approved admin users can rename folders."}

    if folder_id <= 0:
        db.close()
        return {"success": False, "message": "The Unfiled Files folder cannot be renamed."}

    if not name:
        db.close()
        return {"success": False, "message": "Please enter a folder name."}

    if len(name) > 80:
        db.close()
        return {"success": False, "message": "Folder name must be 80 characters or less."}

    folder = db.query(SharedSalesFolder).filter(
        SharedSalesFolder.id == folder_id
    ).first()

    if not folder:
        db.close()
        return {"success": False, "message": "Folder not found."}

    duplicate = db.query(SharedSalesFolder).filter(
        func.lower(SharedSalesFolder.name) == name.lower(),
        SharedSalesFolder.id != folder_id
    ).first()

    if duplicate:
        db.close()
        return {"success": False, "message": "A folder with this name already exists."}

    try:
        folder.name = name
        db.commit()
        result = {
            "success": True,
            "message": "Folder renamed.",
            "folder": {
                "id": folder.id,
                "name": folder.name,
                "created_by": folder.created_by or "",
                "created_at": folder.created_at or ""
            }
        }
        db.close()
        return result
    except Exception as e:
        db.rollback()
        db.close()
        return {"success": False, "message": "Rename failed: " + str(e)}


@app.post("/shared-sales-folders/{folder_id}/delete")
def delete_shared_sales_folder(folder_id: int, data: dict):
    db = SessionLocal()
    username = str(data.get("username", "") or "").strip()

    if not can_manage_shared_files(db, username):
        db.close()
        return {"success": False, "message": "Only BILL and approved admin users can delete folders."}

    if folder_id <= 0:
        db.close()
        return {"success": False, "message": "The Unfiled Files folder cannot be deleted."}

    folder = db.query(SharedSalesFolder).filter(
        SharedSalesFolder.id == folder_id
    ).first()

    if not folder:
        db.close()
        return {"success": False, "message": "Folder not found."}

    try:
        files = db.query(SharedSalesFile).filter(
            SharedSalesFile.folder_id == folder_id
        ).all()

        deleted_file_count = len(files)

        for shared_file in files:
            db.delete(shared_file)

        db.delete(folder)
        db.commit()
        db.close()

        folder_dir = Path("static") / "uploads" / "shared_sales_files" / f"folder_{folder_id}"
        if folder_dir.exists() and folder_dir.is_dir():
            shutil.rmtree(folder_dir, ignore_errors=True)

        return {
            "success": True,
            "message": "Folder deleted.",
            "deleted_file_count": deleted_file_count
        }
    except Exception as e:
        db.rollback()
        db.close()
        return {"success": False, "message": "Folder deletion failed: " + str(e)}


@app.get("/shared-sales-files/{username}")
def get_shared_sales_files(username: str, folder_id: int = None):
    db = SessionLocal()

    if not can_access_shared_files(db, username):
        db.close()
        return {"success": False, "message": "Only sales users and admin can access shared files."}

    query = db.query(SharedSalesFile)

    if folder_id is not None:
        query = query.filter(SharedSalesFile.folder_id == folder_id)

    files = query.order_by(SharedSalesFile.id.desc()).all()

    result = {
        "success": True,
        "can_manage": can_manage_shared_files(db, username),
        "files": [serialize_shared_file(shared_file) for shared_file in files],
        "folders": get_shared_sales_folders(db)
    }
    db.close()
    return result


@app.post("/shared-sales-files/upload-chunk")
def upload_shared_sales_file_chunk(data: dict):
    db = SessionLocal()
    username = str(data.get("username", "") or "").strip()
    upload_id = str(data.get("upload_id", "") or "").strip()
    filename = Path(str(data.get("filename", "") or "")).name.strip()

    try:
        folder_id = int(data.get("folder_id", 0))
    except Exception:
        db.close()
        return {"success": False, "message": "Please select a valid folder."}

    try:
        chunk_index = int(data.get("chunk_index", 0))
        total_chunks = int(data.get("total_chunks", 1))
    except Exception:
        db.close()
        return {"success": False, "message": "Invalid chunk information."}

    if not can_manage_shared_files(db, username):
        db.close()
        return {"success": False, "message": "Only BILL and approved admin users can upload shared files."}

    if not upload_id or not filename:
        db.close()
        return {"success": False, "message": "Missing upload id or file name."}

    folder = db.query(SharedSalesFolder).filter(
        SharedSalesFolder.id == folder_id
    ).first()

    if not folder:
        db.close()
        return {"success": False, "message": "Please create or select a folder before uploading."}

    if chunk_index < 0 or total_chunks <= 0 or chunk_index >= total_chunks:
        db.close()
        return {"success": False, "message": "Invalid chunk index."}

    extension = safe_shared_file_extension(filename)

    if not extension:
        db.close()
        return {
            "success": False,
            "message": "Unsupported file type. Upload images, PDF, Word, Excel, PowerPoint, ZIP, RAR, TXT, CSV or video files."
        }

    chunk_data = str(data.get("chunk_data", "") or "")
    if "," in chunk_data:
        chunk_data = chunk_data.split(",", 1)[1]

    try:
        chunk_bytes = base64.b64decode(chunk_data)
    except Exception:
        db.close()
        return {"success": False, "message": "Invalid chunk data."}

    safe_upload_id = re.sub(r"[^a-zA-Z0-9_-]", "", upload_id)[:80]
    if not safe_upload_id:
        db.close()
        return {"success": False, "message": "Invalid upload id."}

    root_dir = Path("static") / "uploads" / "shared_sales_files"
    tmp_dir = root_dir / "_chunks" / safe_upload_id
    final_dir = root_dir / f"folder_{folder.id}"

    try:
        tmp_dir.mkdir(parents=True, exist_ok=True)
        final_dir.mkdir(parents=True, exist_ok=True)
        (tmp_dir / f"{chunk_index:06d}.part").write_bytes(chunk_bytes)

        if chunk_index < total_chunks - 1:
            db.close()
            return {
                "success": True,
                "message": f"Chunk {chunk_index + 1}/{total_chunks} uploaded.",
                "complete": False
            }

        missing_chunks = [
            index for index in range(total_chunks)
            if not (tmp_dir / f"{index:06d}.part").exists()
        ]
        if missing_chunks:
            db.close()
            return {"success": False, "message": "Missing upload chunks."}

        final_name = datetime.now().strftime("%Y%m%d%H%M%S") + "_" + uuid.uuid4().hex[:10] + extension
        final_path = final_dir / final_name

        with final_path.open("wb") as output_file:
            for index in range(total_chunks):
                output_file.write((tmp_dir / f"{index:06d}.part").read_bytes())

        shutil.rmtree(tmp_dir, ignore_errors=True)

        shared_file = SharedSalesFile(
            folder_id=folder.id,
            original_name=filename[:255],
            file_path="/" + str(final_path).replace("\\", "/"),
            file_type=get_shared_file_type(filename),
            file_size=final_path.stat().st_size,
            uploaded_by=username,
            uploaded_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        db.add(shared_file)
        db.commit()

        result = {
            "success": True,
            "message": "Shared file uploaded successfully.",
            "complete": True,
            "file": serialize_shared_file(shared_file)
        }
        db.close()
        return result
    except Exception as e:
        db.rollback()
        db.close()
        return {"success": False, "message": "Upload failed: " + str(e)}


@app.post("/shared-sales-files/{file_id}/delete")
def delete_shared_sales_file(file_id: int, data: dict):
    db = SessionLocal()
    username = str(data.get("username", "") or "").strip()

    if not can_manage_shared_files(db, username):
        db.close()
        return {"success": False, "message": "Only BILL and approved admin users can delete shared files."}

    shared_file = db.query(SharedSalesFile).filter(
        SharedSalesFile.id == file_id
    ).first()

    if not shared_file:
        db.close()
        return {"success": False, "message": "Shared file not found."}

    local_path = Path("." + str(shared_file.file_path or ""))

    try:
        if local_path.exists() and local_path.is_file():
            local_path.unlink()
        db.delete(shared_file)
        db.commit()
        db.close()
        return {"success": True, "message": "Shared file deleted."}
    except Exception as e:
        db.rollback()
        db.close()
        return {"success": False, "message": "Delete failed: " + str(e)}


# =========================
# ORDERS
# =========================

def can_bill_delete_orders(db, requester_username):
    normalized_username = str(requester_username or "").strip().lower()

    if normalized_username != "bill":
        return False

    bill_user = db.query(User).filter(
        func.lower(User.username) == "bill"
    ).first()

    return bool(
        bill_user and
        bill_user.role == "sales" and
        bill_user.approval_status == "Approved"
    )


@app.post("/create-order")
def create_order(data: dict):
    db = SessionLocal()

    order_username = (
        data.get("order_username") or
        data.get("customer_username") or
        data.get("username")
    )

    if not order_username:
        db.close()
        return {
            "success": False,
            "message": "Customer username is required."
        }

    customer = db.query(User).filter(
        User.username == order_username
    ).first()

    assigned_sales = "BILL"

    if customer and customer.assigned_sales:
        assigned_sales = customer.assigned_sales

    # Sales users can create an order for a customer.
    # In that case, the order still belongs to the customer,
    # but sales_username is the current sales account.
    if data.get("sales_username"):
        assigned_sales = data.get("sales_username")

    items = data.get("items", [])

    for item in items:
        if isinstance(item, dict):
            item["item_status"] = "Pending"
            item["item_return_comment"] = ""

    order = Order(
        username=order_username,
        sales_username=assigned_sales,
        items=json.dumps(items),
        total=data["total"],
        status="Pending",
        created_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        return_comment=""
    )

    db.add(order)
    db.commit()
    db.close()

    return {
        "success": True,
        "message": "Order created successfully!"
    }


@app.get("/sales/orders/{sales_username}")
def get_sales_orders(sales_username: str):
    db = SessionLocal()

    orders = db.query(Order).filter(
        Order.sales_username == sales_username
    ).all()

    result = []

    for order in orders:
        items = normalize_order_items(order)

        result.append({
            "id": order.id,
            "username": order.username,
            "customer": order.username,
            "sales_username": order.sales_username,
            "total": order.total,
            "status": order.status,
            "created_at": order.created_at,
            "return_comment": order.return_comment,
            "items": json.dumps(items),
            "pi": serialize_pi_state(db, order)
        })

    db.close()
    return result


@app.post("/sales/orders/{order_id}/items/{item_index}/confirm")
def confirm_order_item(order_id: int, item_index: int):
    db = SessionLocal()

    order = db.query(Order).filter(
        Order.id == order_id
    ).first()

    if not order:
        db.close()
        return {
            "success": False,
            "message": "Order not found."
        }

    items = normalize_order_items(order)

    if item_index < 0 or item_index >= len(items):
        db.close()
        return {
            "success": False,
            "message": "Product item not found."
        }

    items[item_index]["item_status"] = "Confirmed"
    items[item_index]["item_return_comment"] = ""

    recalculate_order_status(order, items)

    db.commit()
    db.close()

    return {
        "success": True,
        "message": "Product confirmed successfully!"
    }


@app.post("/sales/orders/{order_id}/items/{item_index}/return")
def return_order_item(order_id: int, item_index: int, data: dict):
    db = SessionLocal()

    order = db.query(Order).filter(
        Order.id == order_id
    ).first()

    if not order:
        db.close()
        return {
            "success": False,
            "message": "Order not found."
        }

    return_comment = data.get("return_comment", "").strip()

    if not return_comment:
        db.close()
        return {
            "success": False,
            "message": "Please enter a return comment first."
        }

    items = normalize_order_items(order)

    if item_index < 0 or item_index >= len(items):
        db.close()
        return {
            "success": False,
            "message": "Product item not found."
        }

    items[item_index]["item_status"] = "Returned"
    items[item_index]["item_return_comment"] = return_comment

    recalculate_order_status(order, items)

    db.commit()
    db.close()

    return {
        "success": True,
        "message": "Product returned successfully!"
    }


@app.post("/sales/orders/{order_id}/confirm")
def confirm_order(order_id: int):
    db = SessionLocal()

    order = db.query(Order).filter(
        Order.id == order_id
    ).first()

    if not order:
        db.close()
        return {
            "success": False,
            "message": "Order not found"
        }

    items = normalize_order_items(order)

    for item in items:
        if isinstance(item, dict):
            item["item_status"] = "Confirmed"
            item["item_return_comment"] = ""

    recalculate_order_status(order, items)

    db.commit()
    db.close()

    return {
        "success": True,
        "message": "Order confirmed successfully!"
    }


@app.post("/sales/orders/{order_id}/reject")
@app.post("/sales/orders/{order_id}/return")
def return_order(order_id: int, data: dict):
    db = SessionLocal()

    order = db.query(Order).filter(
        Order.id == order_id
    ).first()

    if not order:
        db.close()
        return {
            "success": False,
            "message": "Order not found"
        }

    return_comment = data.get("return_comment", "").strip()

    if not return_comment:
        db.close()
        return {
            "success": False,
            "message": "Please enter a return comment first."
        }

    items = normalize_order_items(order)

    for item in items:
        if isinstance(item, dict):
            item["item_status"] = "Returned"
            item["item_return_comment"] = return_comment

    recalculate_order_status(order, items)

    db.commit()
    db.close()

    return {
        "success": True,
        "message": "Order returned successfully!"
    }


@app.post("/sales/orders/{order_id}/delete")
def delete_sales_order(order_id: int, data: dict):
    db = SessionLocal()
    requester_username = str(
        data.get("requester_username", "") or ""
    ).strip()

    if not can_bill_delete_orders(db, requester_username):
        db.close()
        return {
            "success": False,
            "message": "Only BILL can delete sales orders."
        }

    order = db.query(Order).filter(
        Order.id == order_id
    ).first()

    if not order:
        db.close()
        return {
            "success": False,
            "message": "Order not found."
        }

    if str(order.sales_username or "").strip().lower() != "bill":
        db.close()
        return {
            "success": False,
            "message": "BILL can only delete orders assigned to BILL."
        }

    try:
        pi_records = db.query(ProformaInvoice).filter(
            ProformaInvoice.order_id == order.id
        ).all()

        pi_histories = db.query(PIHistory).filter(
            PIHistory.order_id == order.id
        ).all()

        pi_file_paths = {
            str(record.pdf_path or "").strip()
            for record in [*pi_records, *pi_histories]
            if str(record.pdf_path or "").strip()
        }

        for history in pi_histories:
            db.delete(history)

        for pi in pi_records:
            db.delete(pi)

        db.delete(order)
        db.commit()
        db.close()

        pi_root = (Path("static") / "uploads" / "pi").resolve()

        for pdf_path in pi_file_paths:
            if not pdf_path.startswith("/static/uploads/pi/"):
                continue

            local_path = (Path(".") / pdf_path.lstrip("/")).resolve()

            if pi_root not in local_path.parents:
                continue

            try:
                if local_path.exists() and local_path.is_file():
                    local_path.unlink()
            except OSError as file_error:
                print(
                    f"Order {order_id} deleted, but PI file cleanup skipped:",
                    file_error
                )

        return {
            "success": True,
            "message": f"Order #{order_id} has been deleted by BILL."
        }

    except Exception as error:
        db.rollback()
        db.close()
        return {
            "success": False,
            "message": "Delete order failed: " + str(error)
        }


@app.get("/customer/info/{username}")
def get_customer_info(username: str):
    db = SessionLocal()

    user = db.query(User).filter(
        User.username == username
    ).first()

    if not user:
        db.close()
        return {
            "success": False,
            "message": "User not found."
        }

    result = {
        "success": True,
        "username": user.username,
        "role": user.role,
        "account_type": user.account_type,
        "company_name": user.company_name,
        "email": user.email,
        "assigned_sales": user.assigned_sales,
        "approval_status": user.approval_status,
        "customer_level": user.customer_level
    }

    db.close()
    return result


@app.get("/customer/orders/{username}")
def get_customer_orders(username: str):
    db = SessionLocal()

    orders = db.query(Order).filter(
        Order.username == username
    ).all()

    result = []

    for order in orders:
        items = normalize_order_items(order)

        result.append({
            "id": order.id,
            "status": order.status,
            "total": order.total,
            "created_at": order.created_at,
            "return_comment": order.return_comment,
            "items": json.dumps(items),
            "pi": serialize_pi_state(db, order)
        })

    db.close()
    return result


@app.post("/customer/orders/{order_id}/delete")
def delete_customer_order(order_id: int, data: dict):
    db = SessionLocal()

    username = data.get("username", "").strip()

    order = db.query(Order).filter(
        Order.id == order_id
    ).first()

    if not order:
        db.close()
        return {
            "success": False,
            "message": "Order not found."
        }

    if order.username != username:
        db.close()
        return {
            "success": False,
            "message": "You can only delete your own orders."
        }

    existing_pi = db.query(ProformaInvoice).filter(
        ProformaInvoice.order_id == order.id
    ).first()

    if existing_pi:
        db.close()
        return {
            "success": False,
            "message": "Orders with PI history cannot be deleted."
        }

    db.delete(order)
    db.commit()
    db.close()

    return {
        "success": True,
        "message": f"Order #{order_id} has been deleted."
    }


# =========================
# PROFORMA INVOICE WORKFLOW
# =========================

@app.post("/sales/orders/{order_id}/pi/send")
def sales_send_pi(order_id: int, data: dict = None):
    db = SessionLocal()

    if data is None:
        data = {}

    order = db.query(Order).filter(
        Order.id == order_id
    ).with_for_update().first()

    if not order:
        db.close()
        return {
            "success": False,
            "message": "Order not found."
        }

    if not order_all_items_confirmed(order):
        db.close()
        return {
            "success": False,
            "message": "PI can only be sent after all products are confirmed."
        }

    sales_username = data.get("sales_username", "").strip() or order.sales_username or ""
    sales_user = None

    if sales_username:
        sales_user = db.query(User).filter(
            User.username == sales_username
        ).first()

    customer = db.query(User).filter(
        User.username == order.username
    ).first()

    latest_pi = get_latest_pi(db, order.id)
    next_version = 1

    if latest_pi:
        latest_status = str(latest_pi.status or "").strip().lower()

        if latest_status != "sent back":
            db.close()

            if latest_status == "received":
                return {
                    "success": False,
                    "message": "The customer has already received this PI."
                }

            return {
                "success": False,
                "message": "Please wait for the customer to receive or send back the current PI."
            }

        next_version = int(latest_pi.version or 1) + 1

    pi_no = f"PI-{datetime.now().strftime('%Y%m%d')}-{order.id:04d}"

    status = "Sent" if next_version == 1 else "Revised Sent"

    try:
        pdf_url = generate_pi_pdf_file(
            order=order,
            customer=customer,
            sales_user=sales_user,
            pi_no=pi_no,
            version=next_version
        )
    except Exception as e:
        db.close()
        return {
            "success": False,
            "message": "PI PDF generation failed: " + str(e)
        }

    pi = ProformaInvoice(
        order_id=order.id,
        pi_no=pi_no,
        pdf_path=pdf_url,
        status=status,
        version=next_version,
        sent_by=sales_username,
        sent_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        received_at="",
        customer_message=""
    )

    db.add(pi)
    db.flush()

    action = "Sales Sent PI" if next_version == 1 else "Sales Sent Revised PI"

    add_pi_history(
        db=db,
        order_id=order.id,
        pi_id=pi.id,
        action=action,
        message=f"{action} version {next_version}.",
        created_by=sales_username or "Sales",
        pdf_path=pdf_url
    )

    db.commit()

    result = {
        "success": True,
        "message": f"PI version {next_version} has been sent.",
        "pi": serialize_pi(pi),
        "pdf_url": pdf_url
    }

    db.close()
    return result


@app.get("/orders/{order_id}/pi")
def get_order_pi(order_id: int):
    db = SessionLocal()

    order = db.query(Order).filter(
        Order.id == order_id
    ).first()

    if not order:
        db.close()
        return {
            "success": False,
            "message": "Order not found."
        }

    result = {
        "success": True,
        "pi": serialize_pi_state(db, order)
    }

    db.close()
    return result


@app.post("/customer/orders/{order_id}/pi/receive")
def customer_receive_pi(order_id: int, data: dict):
    db = SessionLocal()

    username = data.get("username", "").strip()

    order = db.query(Order).filter(
        Order.id == order_id
    ).with_for_update().first()

    if not order:
        db.close()
        return {
            "success": False,
            "message": "Order not found."
        }

    if order.username != username:
        db.close()
        return {
            "success": False,
            "message": "You can only receive PI for your own order."
        }

    pi = get_latest_pi(db, order.id)

    if not pi:
        db.close()
        return {
            "success": False,
            "message": "No PI has been sent for this order yet."
        }

    if pi.status == "Received":
        db.close()
        return {
            "success": False,
            "message": "This PI has already been received."
        }

    if not pi_waiting_for_customer(pi):
        db.close()
        return {
            "success": False,
            "message": "This PI is not waiting for customer confirmation."
        }

    pi.status = "Received"
    pi.received_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    add_pi_history(
        db=db,
        order_id=order.id,
        pi_id=pi.id,
        action="Customer Received PI",
        message="Customer confirmed PI received.",
        created_by=username,
        pdf_path=pi.pdf_path
    )

    db.commit()

    result = {
        "success": True,
        "message": "PI received successfully.",
        "pi": serialize_pi(pi)
    }

    db.close()
    return result


@app.post("/customer/orders/{order_id}/pi/send-back")
def customer_send_back_pi(order_id: int, data: dict):
    db = SessionLocal()

    username = data.get("username", "").strip()
    message = data.get("message", "").strip()

    if not message:
        db.close()
        return {
            "success": False,
            "message": "Please enter a sent-back reason."
        }

    order = db.query(Order).filter(
        Order.id == order_id
    ).with_for_update().first()

    if not order:
        db.close()
        return {
            "success": False,
            "message": "Order not found."
        }

    if order.username != username:
        db.close()
        return {
            "success": False,
            "message": "You can only send back PI for your own order."
        }

    pi = get_latest_pi(db, order.id)

    if not pi:
        db.close()
        return {
            "success": False,
            "message": "No PI has been sent for this order yet."
        }

    if pi.status == "Received":
        db.close()
        return {
            "success": False,
            "message": "Received PI cannot be sent back."
        }

    if not pi_waiting_for_customer(pi):
        db.close()
        return {
            "success": False,
            "message": "This PI has already been sent back or is no longer awaiting a response."
        }

    pi.status = "Sent Back"
    pi.customer_message = message

    add_pi_history(
        db=db,
        order_id=order.id,
        pi_id=pi.id,
        action="Customer Sent Back PI",
        message=message,
        created_by=username,
        pdf_path=pi.pdf_path
    )

    db.commit()

    result = {
        "success": True,
        "message": "PI has been sent back to sales.",
        "pi": serialize_pi(pi)
    }

    db.close()
    return result



# =========================
# PDF QUOTATION
# =========================

@app.post("/cart/pdf")
async def generate_cart_pdf(data: dict):

    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        Image
    )
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.utils import ImageReader
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from pathlib import Path as LocalPath
    from datetime import timedelta

    items = data.get("items", [])

    customer_name = data.get("customerName", "")
    customer_company = data.get("customerCompany", "")
    customer_email = data.get("customerEmail", "")
    username = data.get("username", "")
    sales_username_from_cart = data.get("salesUsername", "")

    try:
        valid_days = int(data.get("validDays", 30))
    except Exception:
        valid_days = 30

    if valid_days <= 0:
        valid_days = 30

    quote_date = datetime.now()
    valid_until_date = quote_date + timedelta(days=valid_days)

    quotation_no = "BQ-" + quote_date.strftime("%Y%m%d%H%M")

    # Get assigned sales contact.
    # Normal customer flow: use the customer's assigned_sales.
    # Sales-mode flow: cart.html sends salesUsername, so use that sales account first.
    sales_name = sales_username_from_cart or ""
    sales_email = ""

    db = SessionLocal()

    try:
        if not sales_name:
            customer_user = db.query(User).filter(
                User.username == username
            ).first()

            if customer_user:
                sales_name = customer_user.assigned_sales or ""

        if sales_name:
            sales_user = db.query(User).filter(
                User.username == sales_name
            ).first()

            if sales_user:
                sales_name = sales_user.username or sales_name
                sales_email = sales_user.email or ""

    finally:
        db.close()

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=24,
        leftMargin=24,
        topMargin=22,
        bottomMargin=22
    )

    elements = []
    styles = getSampleStyleSheet()

    brand_blue = colors.HexColor("#1f3c88")
    action_blue = colors.HexColor("#2563eb")
    light_blue = colors.HexColor("#eef4ff")
    soft_gray = colors.HexColor("#f8fafc")
    border_gray = colors.HexColor("#dbe3ef")
    dark_text = colors.HexColor("#111827")
    muted_text = colors.HexColor("#4b5563")

    title_style = ParagraphStyle(
        "BridgeTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=22,
        leading=26,
        textColor=brand_blue,
        alignment=TA_RIGHT,
        spaceAfter=0
    )

    subtitle_style = ParagraphStyle(
        "BridgeSubtitle",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        textColor=muted_text,
        alignment=TA_RIGHT
    )

    section_title_style = ParagraphStyle(
        "SectionTitle",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=brand_blue,
        spaceAfter=6
    )

    normal_style = ParagraphStyle(
        "NormalClean",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        textColor=dark_text
    )

    small_style = ParagraphStyle(
        "SmallText",
        parent=styles["Normal"],
        fontSize=7.8,
        leading=10,
        textColor=dark_text
    )

    small_center_style = ParagraphStyle(
        "SmallCenterText",
        parent=small_style,
        alignment=TA_CENTER
    )

    small_right_style = ParagraphStyle(
        "SmallRightText",
        parent=small_style,
        alignment=TA_RIGHT
    )

    note_style = ParagraphStyle(
        "NoteText",
        parent=styles["Normal"],
        fontSize=8,
        leading=11,
        textColor=muted_text
    )

    def escape_pdf_text(value):
        if value is None:
            value = ""

        text = str(value)

        return (
            text
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace("\n", "<br/>")
        )

    def make_text(value, center=False, right=False):
        if center:
            return Paragraph(escape_pdf_text(value), small_center_style)

        if right:
            return Paragraph(escape_pdf_text(value), small_right_style)

        return Paragraph(escape_pdf_text(value), small_style)

    def money(value):
        try:
            return "${:,.2f}".format(float(value or 0))
        except Exception:
            return "$0.00"

    def safe_number(value, default=0):
        try:
            return float(value or default)
        except Exception:
            return float(default)

    def make_info_line(label, value):
        return f"<b>{escape_pdf_text(label)}:</b> {escape_pdf_text(value)}"

    def make_product_image(image_url):
        if not image_url:
            return make_text("No Image", center=True)

        image_path = str(image_url)

        if image_path.startswith("/"):
            image_path = "." + image_path

        local_path = LocalPath(image_path)

        if not local_path.exists():
            return make_text("No Image", center=True)

        try:
            reader = ImageReader(str(local_path))
            img_width, img_height = reader.getSize()

            max_width = 44
            max_height = 44

            scale = min(max_width / img_width, max_height / img_height)

            product_image = Image(
                str(local_path),
                width=img_width * scale,
                height=img_height * scale
            )

            product_image.hAlign = "CENTER"

            return product_image

        except Exception:
            return make_text("No Image", center=True)

    # Header
    logo_path = LocalPath("static/image/bg2.jpg")

    if logo_path.exists():
        logo = Image(
            str(logo_path),
            width=230,
            height=40
        )
        logo.hAlign = "LEFT"
        logo_cell = logo
    else:
        logo_cell = Paragraph(
            "<b>BRIDGE QUOTATION</b>",
            ParagraphStyle(
                "TextLogo",
                parent=styles["Normal"],
                fontName="Helvetica-Bold",
                fontSize=16,
                textColor=brand_blue
            )
        )

    title_block = [
        Paragraph("QUOTATION", title_style),
        Paragraph("Product selection and quotation platform", subtitle_style)
    ]

    header_table = Table(
        [[logo_cell, title_block]],
        colWidths=[250, 285]
    )

    header_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -1), 1.2, brand_blue),
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 14))

    # Quote and contact information
    quote_info = Paragraph(
        "<b>Quotation Details</b><br/><br/>"
        + make_info_line("Quotation No", quotation_no)
        + "<br/>"
        + make_info_line("Website", "bridgequotation.com")
        + "<br/>"
        + make_info_line("Date", quote_date.strftime("%Y-%m-%d"))
        + "<br/>"
        + make_info_line("Valid Until", valid_until_date.strftime("%Y-%m-%d"))
        + f" ({valid_days} days)",
        normal_style
    )

    customer_info = Paragraph(
        "<b>Customer Information</b><br/><br/>"
        + make_info_line("Customer", customer_name)
        + "<br/>"
        + make_info_line("Company", customer_company)
        + "<br/>"
        + make_info_line("Email", customer_email),
        normal_style
    )

    sales_info = Paragraph(
        "<b>Sales Contact</b><br/><br/>"
        + make_info_line("Name", sales_name or "Not assigned")
        + "<br/>"
        + make_info_line("Email", sales_email or ""),
        normal_style
    )

    info_table = Table(
        [[quote_info, customer_info, sales_info]],
        colWidths=[178, 178, 178]
    )

    info_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), soft_gray),
        ("BOX", (0, 0), (-1, -1), 0.8, border_gray),
        ("INNERGRID", (0, 0), (-1, -1), 0.6, border_gray),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))

    elements.append(info_table)
    elements.append(Spacer(1, 16))

    elements.append(Paragraph("Product Quotation", section_title_style))

    table_data = [
        [
            make_text("Image", center=True),
            make_text("Product", center=True),
            make_text("MOQ", center=True),
            make_text("Material", center=True),
            make_text("Size", center=True),
            make_text("Packing", center=True),
            make_text("Qty", center=True),
            make_text("Unit Price", center=True),
            make_text("Amount", center=True)
        ]
    ]

    grand_total = 0
    total_quantity = 0

    for item in items:

        qty = int(safe_number(item.get("quantity", 0)))
        price = safe_number(item.get("price", 0))
        total = qty * price

        total_quantity += qty
        grand_total += total

        product_image = make_product_image(item.get("image", ""))

        product_name = item.get("name", "Product")
        requirement = item.get("customer_requirement", "")

        product_lines = f"<b>{escape_pdf_text(product_name)}</b>"

        if requirement:
            product_lines += (
                "<br/><font size='6.8' color='#4b5563'>"
                + "Req: "
                + escape_pdf_text(requirement)
                + "</font>"
            )

        table_data.append([
            product_image,
            Paragraph(product_lines, small_style),
            make_text(item.get("moq", ""), center=True),
            make_text(item.get("material", ""), center=True),
            make_text(item.get("size", ""), center=True),
            make_text(item.get("packing", ""), center=True),
            make_text(f"{qty:,}", center=True),
            make_text(money(price), right=True),
            make_text(money(total), right=True)
        ])

    table_data.append([
        "",
        Paragraph("<b>Grand Total</b>", small_right_style),
        "",
        "",
        "",
        "",
        make_text(f"{total_quantity:,}", center=True),
        "",
        Paragraph(f"<b>{money(grand_total)}</b>", small_right_style)
    ])

    product_table = Table(
        table_data,
        colWidths=[48, 110, 38, 52, 48, 62, 42, 58, 70],
        repeatRows=1
    )

    product_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand_blue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

        ("BACKGROUND", (0, 1), (-1, -2), colors.white),
        ("BACKGROUND", (0, -1), (-1, -1), light_blue),

        ("BOX", (0, 0), (-1, -1), 0.8, border_gray),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, border_gray),

        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("ALIGN", (2, 1), (6, -1), "CENTER"),
        ("ALIGN", (7, 1), (8, -1), "RIGHT"),

        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(product_table)
    elements.append(Spacer(1, 18))

    notes_title = Paragraph("Notes", section_title_style)

    notes = Paragraph(
        "1. Prices are valid within the quotation period shown above.<br/>"
        "2. Final price is subject to confirmed quantity, packing details and product availability.<br/>"
        "3. Customized packing, artwork, shipping marks and other special requirements should be confirmed before production.<br/>"
        "4. This quotation is generated from bridgequotation.com.",
        note_style
    )

    notes_table = Table(
        [[notes_title], [notes]],
        colWidths=[535]
    )

    notes_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), soft_gray),
        ("BOX", (0, 0), (-1, -1), 0.8, border_gray),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    elements.append(notes_table)
    elements.append(Spacer(1, 14))

    footer = Paragraph(
        "<font size=8 color='#6b7280'>Thank you for your inquiry. Generated by bridgequotation.com.</font>",
        styles["Normal"]
    )

    elements.append(footer)

    doc.build(elements)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": "attachment; filename=bridge_quotation.pdf"
        }
    )


# =========================
# DEBUG TOOLS
# =========================

@app.get("/debug/users")
def debug_users():
    db = SessionLocal()

    users = db.query(User).all()

    result = [
        {
            "id": u.id,
            "username": u.username,
            "role": u.role,
            "approval_status": u.approval_status,
            "company_name": u.company_name,
            "email": u.email
        }
        for u in users
    ]

    db.close()

    return result


@app.get("/debug/make-sales/{username}")
def make_sales(username: str):
    db = SessionLocal()

    user = db.query(User).filter(
        User.username == username
    ).first()

    if not user:
        db.close()
        return {
            "success": False,
            "message": "User not found"
        }

    user.role = "sales"
    user.account_type = "employee"
    user.approval_status = "Approved"
    user.assigned_sales = ""

    db.commit()
    db.close()

    return {
        "success": True,
        "message": f"{username} is now a sales user"
    }


@app.get("/debug/delete-order/{order_id}")
def delete_order(order_id: int):
    return {
        "success": False,
        "message": "Debug order deletion is disabled. Use the BILL Sales Dashboard."
    }
